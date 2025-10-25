import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

# ייבוא מידע גרסה
try:
    from version import __version__, __app_name__, get_version_string
except ImportError:
    __version__ = "1.0.0"
    __app_name__ = "מעקב אימונים"
    def get_version_string():
        return f"{__app_name__} v{__version__}"

# Optional dependencies: import lazily and tolerate absence so module can be
# imported in environments missing optional packages (e.g., CI/test).
try:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except Exception:
    Workbook = None  # type: ignore
    LineChart = Reference = None  # type: ignore
    Alignment = Font = PatternFill = None  # type: ignore
    get_column_letter = None  # type: ignore
    _HAS_OPENPYXL = False

# Matplotlib / Qt canvas may be optional in headless/test environments.
os.environ.setdefault('QT_API', 'pyside6')
try:
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    import matplotlib.dates as mdates
    _HAS_MPL = True
except Exception:
    FigureCanvas = None  # type: ignore
    Figure = None  # type: ignore
    mdates = None  # type: ignore
    _HAS_MPL = False
try:
    from PySide6.QtCore import QDate, QEvent, QSize, Qt, QTimer, QRectF, QPointF
    from PySide6.QtGui import (
        QAction,
        QColor,
        QDoubleValidator,
        QFont,
        QIntValidator,
        QKeySequence,
        QPainter,
        QPainterPath,
        QPen,
        QPixmap,
        QShortcut,
        QValidator,
    )
    from PySide6.QtWidgets import (
        QApplication,
        QButtonGroup,
        QCalendarWidget,
        QDialog,
        QDialogButtonBox,
        QFileDialog,
        QFrame,
        QGridLayout,
        QHBoxLayout,
        QInputDialog,
        QLabel,
        QLineEdit,
        QListWidget,
        QListWidgetItem,
        QMainWindow,
        QMenu,
        QMessageBox,
        QPushButton,
        QRadioButton,
        QSizePolicy,
        QStatusBar,
        QTableWidget,
        QTableWidgetItem,
        QTabWidget,
        QToolBar,
        QVBoxLayout,
        QWidget,
    )
    _HAS_QT = True
except Exception:
    # If PySide6 isn't installed, set a flag and define minimal stand-ins for type checkers.
    _HAS_QT = False
    # Define placeholders so static analysis of the file can continue in limited fashion.
    QDate = QEvent = QSize = Qt = object
    QAction = QColor = QDoubleValidator = QIntValidator = QKeySequence = QShortcut = QValidator = object
    QApplication = QButtonGroup = QCalendarWidget = QDialog = QDialogButtonBox = QFileDialog = QFrame = QGridLayout = QHBoxLayout = QInputDialog = QLabel = QLineEdit = QListWidget = QListWidgetItem = QMainWindow = QMenu = QMessageBox = QPushButton = QRadioButton = QSizePolicy = QStatusBar = QTableWidget = QTableWidgetItem = QTabWidget = QToolBar = QVBoxLayout = QWidget = object


# טבלה שמאזנת עמודות לרוחב שווה בכל שינוי גודל
class EqualWidthTable(QTableWidget):
    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._equalize_columns()

    def showEvent(self, event):
        super().showEvent(event)
        self._equalize_columns()

    def _equalize_columns(self):
        cols = self.columnCount()
        if cols <= 0:
            return
        width = self.viewport().width()
        col_width = width // cols
        for c in range(cols):
            self.setColumnWidth(c, col_width)


class SummaryTab(QWidget):
    """גיליון סיכום כללי של כל התרגילים"""
    def __init__(self):
        super().__init__()
        self.setContentsMargins(10, 10, 10, 10)
        self._init_ui()
    
    def _init_ui(self):
        """יצירת ממשק המשתמש לגיליון הסיכום"""
        layout = QVBoxLayout()
        layout.setSpacing(20)
        
        # כותרת
        title_label = QLabel("📊 סיכום כללי")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 24pt;
                font-weight: bold;
                color: #2196F3;
                padding: 20px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #E3F2FD, stop:1 #BBDEFB);
                border-radius: 10px;
            }
        """)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # הודעה זמנית
        info_label = QLabel("🚧 גיליון זה בבנייה...\n\nבעתיד יוצגו כאן:\n• סיכום כללי של כל התרגילים\n• גרפי השוואה\n• סטטיסטיקות מתקדמות")
        info_label.setStyleSheet("""
            QLabel {
                font-size: 14pt;
                color: #666;
                padding: 40px;
                background-color: #FAFAFA;
                border: 2px dashed #BDBDBD;
                border-radius: 8px;
            }
        """)
        info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(info_label)
        
        layout.addStretch()
        self.setLayout(layout)


class ExerciseTab(QWidget):
    def __init__(self, exercise_name: str, profile_name: str = None):
        super().__init__()
        self.exercise_name = exercise_name
        self.profile_name = profile_name or "ברירת מחדל"  # פרופיל ברירת מחדל אם לא צוין
        self.setContentsMargins(5, 5, 5, 5)
        self._has_unsaved_changes = False
        # מערכת Undo/Redo
        self._undo_stack = []  # מחסנית של מצבי טבלה קודמים
        self._redo_stack = []  # מחסנית של מצבים לשחזור
        self._max_undo = 5  # מקסימום 5 פעולות
        self._is_restoring = False  # דגל למניעת שמירה בזמן שחזור
        self._init_ui()
        try:
            self.load_state()
        except Exception:
            pass
        # אחרי טעינת המצב, נאפס את דגל השינויים
        self._has_unsaved_changes = False
        # שמירת מצב ראשוני
        self._save_state_to_undo()

    def _show_status(self, message: str, duration: int = 2000):
        """הצגת הודעה בסטטוס בר"""
        window = self.window()
        if isinstance(window, QMainWindow) and window.statusBar():
            window.statusBar().showMessage(message, duration)

    def _init_ui(self):
        layout = QVBoxLayout()

        # טופס הכנסת נתונים
        form = QGridLayout()
        form.setContentsMargins(0, 0, 0, 0)

        # הגדרת שורות טופס ההכנסה
        # תאריך ומשקל
        self.input_weight = QLineEdit()
        self.input_weight.setPlaceholderText("משקל")
        self.input_weight.setValidator(QDoubleValidator(0, 1000, 3))

        # סטים וחזרות
        self.input_sets = QLineEdit()
        self.input_sets.setPlaceholderText("סטים")
        self.input_sets.setValidator(QIntValidator(0, 1000))

        self.input_reps = QLineEdit()
        self.input_reps.setPlaceholderText("חזרות")
        self.input_reps.setValidator(QIntValidator(0, 1000))

        self.input_last_reps = QLineEdit()
        self.input_last_reps.setPlaceholderText("סט אחרון")
        self.input_last_reps.setValidator(QIntValidator(0, 1000))

        # כפתורים: הוסף ומחק
        self.btn_add = QPushButton("הוסף")
        self.btn_pop = QPushButton("מחק אחרון")
        self.btn_delete_row = QPushButton("מחק שורה")
        self.btn_duplicate_row = QPushButton("שכפל שורה")
        self.btn_plot = QPushButton("הצג גרף")
        self.btn_back = QPushButton("חזור לטבלה")
        self.btn_back.hide()
        
        # סגנון מיוחד לכפתורים
        self.btn_plot.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
            }
            QPushButton:hover {
                background-color: #388E3C;
            }
        """)
        
        delete_buttons_style = """
            QPushButton {
                background-color: #f44336;
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
            QPushButton:disabled {
                background-color: #ffcdd2;
            }
        """
        self.btn_pop.setStyleSheet(delete_buttons_style)
        self.btn_delete_row.setStyleSheet(delete_buttons_style)
        
        # עיצוב כפתור שכפול
        duplicate_button_style = """
            QPushButton {
                background-color: #FF9800;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
            QPushButton:disabled {
                background-color: #FFE0B2;
            }
        """
        self.btn_duplicate_row.setStyleSheet(duplicate_button_style)
        
        # התחלת מצב כפתורים - מבוטלים
        self.btn_add.setEnabled(False)
        self.btn_pop.setEnabled(False)
        self.btn_delete_row.setEnabled(False)
        self.btn_duplicate_row.setEnabled(False)

        # יצירת תצוגת סיכום
        summary_layout = QHBoxLayout()
        summary_layout.setSpacing(15)
        
        # עיצוב תוויות הסיכום בקופסאות
        # קופסה כחולה לאימונים
        exercises_style = """
            QLabel {
                font-size: 16pt;
                font-weight: bold;
                color: white;
                padding: 15px 25px;
                border-radius: 8px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2196F3, stop:1 #1976D2);
                border: 2px solid #1565C0;
            }
        """
        
        # קופסה ירוקה למשקל שהרמתי
        weight_style = """
            QLabel {
                font-size: 16pt;
                font-weight: bold;
                color: white;
                padding: 15px 25px;
                border-radius: 8px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #4CAF50, stop:1 #388E3C);
                border: 2px solid #2E7D32;
            }
        """
        
        # קופסה כתומה לממוצע
        avg_style = """
            QLabel {
                font-size: 16pt;
                font-weight: bold;
                color: white;
                padding: 15px 25px;
                border-radius: 8px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #FF9800, stop:1 #F57C00);
                border: 2px solid #E65100;
            }
        """
        
        self.total_exercises_label = QLabel('<div style="text-align: center;">תרגילים<br><span style="font-size: 24pt;">0</span><br><span style="font-size: 32pt;">💪</span></div>')
        self.total_exercises_label.setStyleSheet(exercises_style)
        self.total_exercises_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.total_exercises_label.setMinimumWidth(300)
        self.total_exercises_label.setMaximumWidth(300)
        self.total_exercises_label.setTextFormat(Qt.TextFormat.RichText)
        
        self.total_weight_label = QLabel('<div style="text-align: center;">משקל שהרמתי<br><span style="font-size: 24pt;">0 ק"ג</span><br><span style="font-size: 32pt;">🏋️</span></div>')
        self.total_weight_label.setStyleSheet(weight_style)
        self.total_weight_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.total_weight_label.setMinimumWidth(300)
        self.total_weight_label.setMaximumWidth(300)
        self.total_weight_label.setTextFormat(Qt.TextFormat.RichText)
        
        self.avg_weight_label = QLabel('<div style="text-align: center;">משקל לסט<br><span style="font-size: 24pt;">0 ק"ג</span><br><span style="font-size: 32pt;">📊</span></div>')
        self.avg_weight_label.setStyleSheet(avg_style)
        self.avg_weight_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.avg_weight_label.setMinimumWidth(300)
        self.avg_weight_label.setMaximumWidth(300)
        self.avg_weight_label.setTextFormat(Qt.TextFormat.RichText)
        
        # קופסה סגולה לרמת התקדמות
        progress_style = """
            QLabel {
                font-size: 16pt;
                font-weight: bold;
                color: white;
                padding: 15px 25px;
                border-radius: 8px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #9C27B0, stop:1 #7B1FA2);
                border: 2px solid #6A1B9A;
            }
        """
        
        self.progress_label = QLabel('<div style="text-align: center;">רמה<br><span style="font-size: 20pt;">טירון</span><br><span style="font-size: 32pt;">🌱</span></div>')
        self.progress_label.setStyleSheet(progress_style)
        self.progress_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.progress_label.setMinimumWidth(300)
        self.progress_label.setMaximumWidth(300)
        self.progress_label.setTextFormat(Qt.TextFormat.RichText)
        
        summary_layout.addWidget(self.total_exercises_label)
        summary_layout.addWidget(self.total_weight_label)
        summary_layout.addWidget(self.avg_weight_label)
        summary_layout.addWidget(self.progress_label)
        
        # רשימת שדות קלט
        self._inputs = [
            self.input_weight,
            self.input_sets,
            self.input_reps,
            self.input_last_reps,
        ]
        
        # הוספת שדות לטופס ללא תוויות
        input_layout = QVBoxLayout()
        
        # הגדרת רוחב מקסימלי לשדות הקלט
        for field in self._inputs:
            field.setMaximumWidth(150)
            input_layout.addWidget(field)
        
        # סידור השדות והסיכום בשורה אחת
        inputs_and_summary = QHBoxLayout()
        inputs_and_summary.addLayout(summary_layout)
        inputs_and_summary.addStretch()
        inputs_and_summary.addLayout(input_layout)
        
        form.addLayout(inputs_and_summary, 0, 0)

        # חיבור אירועי שדות קלט
        for inp in self._inputs:
            inp.textChanged.connect(self._update_add_enabled)
            inp.returnPressed.connect(self._try_add_on_enter)
            inp.installEventFilter(self)

        # טבלת נתונים עם עמודות שוות רוחב
        self.table = EqualWidthTable()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["סט אחרון", "חזרות", "סטים", "משקל", "תאריך"])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)  # ביטול עריכה ישירה
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_table_context_menu)
        self.table.cellDoubleClicked.connect(self._edit_date_cell)  # חיבור לאירוע לחיצה כפולה

        # אירועי כפתורים
        self.btn_add.clicked.connect(self.add_entry)
        self.btn_pop.clicked.connect(self.pop_last)
        self.btn_delete_row.clicked.connect(self.delete_selected_row)
        self.btn_duplicate_row.clicked.connect(self.duplicate_selected_row)
        self.btn_plot.clicked.connect(self.plot_selected_exercise)
        self.btn_back.clicked.connect(self.restore_normal_view)
        
        # חיבור לאירוע בחירת שורה בטבלה
        self.table.itemSelectionChanged.connect(self._update_delete_button)
        
        # קיצורי מקלדת למחיקה ושכפול שורה
        delete_shortcut = QShortcut(QKeySequence("Ctrl+E"), self)
        delete_shortcut.activated.connect(self.delete_selected_row)
        
        delete_shortcut_he = QShortcut(QKeySequence("Ctrl+ק"), self)
        delete_shortcut_he.activated.connect(self.delete_selected_row)
        
        duplicate_shortcut = QShortcut(QKeySequence("Ctrl+D"), self)
        duplicate_shortcut.activated.connect(self.duplicate_selected_row)
        
        duplicate_shortcut_he = QShortcut(QKeySequence("Ctrl+ג"), self)
        duplicate_shortcut_he.activated.connect(self.duplicate_selected_row)

        # מסגרת גרף
        self.figure = Figure(figsize=(6, 4))
        self.canvas = FigureCanvas(self.figure)

        # הוספת רכיבים לממשק
        bottom_buttons = QHBoxLayout()
        bottom_buttons.addWidget(self.btn_add)
        bottom_buttons.addWidget(self.btn_pop)
        bottom_buttons.addWidget(self.btn_delete_row)
        bottom_buttons.addWidget(self.btn_duplicate_row)
        bottom_buttons.addWidget(self.btn_plot)
        bottom_buttons.addWidget(self.btn_back)

        self.input_container = QWidget()
        self.input_container.setLayout(form)
        self.input_container.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Maximum)

        layout.addWidget(self.input_container)
        layout.addLayout(bottom_buttons)
        layout.addWidget(self.table)
        layout.addWidget(self.canvas)

        self.setLayout(layout)

    def _update_add_enabled(self):
        weight_ok = self._validate_input(self.input_weight, self.input_weight.text().strip().replace(",", "."))
        sets_ok = self._validate_input(self.input_sets)
        reps_ok = self._validate_input(self.input_reps)
        last_reps_ok = self._validate_input(self.input_last_reps)
        self.btn_add.setEnabled(weight_ok and sets_ok and reps_ok and last_reps_ok)

    def _validate_input(self, widget: Any, text: str = None) -> bool:
        """בדיקת תקינות קלט עבור שדה"""
        v = widget.validator()
        if v is None:
            return False
        test_text = text if text is not None else widget.text()
        res = v.validate(test_text, 0)
        if isinstance(res, tuple) and len(res) > 0:
            return res[0] == QValidator.State.Acceptable
        return False

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Type.KeyPress and obj in self._inputs:
            key = event.key()
            idx = self._inputs.index(obj)
            if key == Qt.Key.Key_Down:
                self._inputs[(idx + 1) % len(self._inputs)].setFocus()
                return True
            if key == Qt.Key.Key_Up:
                self._inputs[(idx - 1) % len(self._inputs)].setFocus()
                return True
        return super().eventFilter(obj, event)

    def _try_add_on_enter(self):
        if self.btn_add.isEnabled():
            self.add_entry()

    def _calculate_total_weight(self):
        """חישוב סך המשקל המצטבר מכל האימונים"""
        total = 0
        for row in range(self.table.rowCount()):
            try:
                items = [self.table.item(row, i) for i in [3, 2, 1, 0]]  # weight, sets, reps, last_reps
                
                # בדיקה מקיפה של תקינות הנתונים
                if not all(isinstance(item, QTableWidgetItem) and item.text() for item in items):
                    continue

                # המרת הערכים למספרים
                weight = float(items[0].text().split()[0].replace(",", "."))
                sets = int(items[1].text())
                reps = int(items[2].text())
                last_reps = int(items[3].text())
                
                # חישוב: (סטים-1 * חזרות * משקל) + (סט אחרון * משקל)
                total += ((sets - 1) * reps * weight) + (last_reps * weight)
            except (ValueError, AttributeError, IndexError):
                continue
        return total

    def _update_summary(self):
        """עדכון תוויות הסיכום"""
        # עדכון מספר התרגילים
        exercises_count = self.table.rowCount()
        self.total_exercises_label.setText(f'<div style="text-align: center;">תרגילים<br><span style="font-size: 24pt;">{exercises_count}</span><br><span style="font-size: 32pt;">💪</span></div>')
        
        # עדכון סך המשקל
        total_weight = self._calculate_total_weight()
        self.total_weight_label.setText(f'<div style="text-align: center;">משקל שהרמתי<br><span style="font-size: 24pt;">{total_weight:,.0f} ק"ג</span><br><span style="font-size: 32pt;">🏋️</span></div>')
        
        # עדכון משקל ממוצע לסט
        if exercises_count > 0:
            avg_weight = total_weight / exercises_count
            self.avg_weight_label.setText(f'<div style="text-align: center;">משקל לסט<br><span style="font-size: 24pt;">{avg_weight:,.0f} ק"ג</span><br><span style="font-size: 32pt;">📊</span></div>')
        else:
            self.avg_weight_label.setText('<div style="text-align: center;">משקל לסט<br><span style="font-size: 24pt;">0 ק"ג</span><br><span style="font-size: 32pt;">📊</span></div>')
        
        # עדכון רמת התקדמות
        self._update_progress_level(exercises_count)
    
    def _update_progress_level(self, exercises_count):
        """עדכון רמת התקדמות על פי מספר האימונים"""
        # הגדרת שלבים
        levels = [
            (0, 10, "טירון", "🌱", 0),
            (10, 30, "מתחיל", "🌿", 1),
            (30, 60, "מתקדם", "🌳", 2),
            (60, 100, "מומחה", "🏆", 3),
            (100, float('inf'), "אגדי", "👑", 4)
        ]
        
        # מציאת הרמה הנוכחית
        current_level = levels[0]
        for level in levels:
            min_val, max_val, name, emoji, level_num = level
            if min_val <= exercises_count < max_val:
                current_level = level
                break
        
        min_val, max_val, level_name, emoji, level_num = current_level
        
        # חישוב אחוזי התקדמות ברמה הנוכחית
        if max_val == float('inf'):
            progress_percent = 100
            next_milestone = "מקסימום!"
        else:
            progress_in_level = exercises_count - min_val
            level_range = max_val - min_val
            progress_percent = (progress_in_level / level_range) * 100
            next_milestone = f"עד {max_val}"
        
        # יצירת פס התקדמות ויזואלי
        # 3 קווים מייצגים את השלבים
        total_levels = 5
        filled_levels = level_num
        
        # יצירת פס עם נקודות
        progress_dots = ""
        for i in range(total_levels):
            if i < filled_levels:
                progress_dots += "●"  # נקודה מלאה
            elif i == filled_levels:
                # נקודה חלקית על פי האחוז
                if progress_percent >= 66:
                    progress_dots += "◉"  # כמעט מלא
                elif progress_percent >= 33:
                    progress_dots += "◔"  # חצי
                else:
                    progress_dots += "○"  # ריק
            else:
                progress_dots += "○"  # נקודה ריקה
            
            if i < total_levels - 1:
                progress_dots += "━"  # קו מחבר
        
        # עדכון התווית
        progress_html = f'''
        <div style="text-align: center;">
            <span style="font-size: 14pt;">רמה</span><br>
            <span style="font-size: 22pt; font-weight: bold;">{level_name}</span><br>
            <span style="font-size: 28pt;">{emoji}</span><br>
            <span style="font-size: 12pt;">{progress_dots}</span><br>
            <span style="font-size: 11pt;">{exercises_count} תרגילים | {next_milestone}</span>
        </div>
        '''
        self.progress_label.setText(progress_html)

    def add_entry(self):
        weight_raw = self.input_weight.text().strip().replace(",", ".")
        sets_raw = self.input_sets.text().strip()
        reps_raw = self.input_reps.text().strip()
        last_reps_raw = self.input_last_reps.text().strip()

        if not (weight_raw and sets_raw and reps_raw and last_reps_raw):
            self._show_status("מלא את כל השדות.")
            return
            
        self._has_unsaved_changes = True

        try:
            weight_val = float(weight_raw)
            weight_str = f"{int(weight_val)}" if weight_val.is_integer() else f"{weight_val:.3f}".rstrip("0").rstrip(".")
            sets_val = int(sets_raw)
            reps_val = int(reps_raw)
            last_reps_val = int(last_reps_raw)
        except ValueError:
            self._show_status("קלט לא תקין.")
            return

        # תאריך
        date_str = datetime.now().strftime("%d/%m/%Y")

        # שמירה למחסנית Undo לפני השינוי
        self._save_state_to_undo()

        # הוספה לטבלה
        row = self.table.rowCount()
        self.table.insertRow(row)

        data = [last_reps_val, reps_val, sets_val, f"{weight_str} Kg", date_str]
        
        for col, value in enumerate(data):
            item = QTableWidgetItem(str(value))
            item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row, col, item)
        
        # עדכון הסיכום
        self._update_summary()

        # ניקוי שדות
        for field in self._inputs:
            field.clear()
        self.input_weight.setFocus()
        self._update_add_enabled()
        self.btn_pop.setEnabled(True)
        self._show_status(f"התווסף: {weight_str} Kg, {sets_val}x{reps_val}")

    def pop_last(self):
        rows = self.table.rowCount()
        if rows > 0:
            # שמירה למחסנית Undo לפני השינוי
            self._save_state_to_undo()
            self.table.removeRow(rows - 1)
            self.btn_pop.setEnabled(self.table.rowCount() > 0)
            self._has_unsaved_changes = True
            self._update_summary()
            self._show_status("נמחק האחרון.")

    def plot_selected_exercise(self):
        # הסתר את האזורים שלא נחוצים בתצוגת גרף
        self.input_container.hide()
        self.table.hide()
        self.btn_add.hide()
        self.btn_pop.hide()
        self.btn_delete_row.hide()
        self.btn_duplicate_row.hide()
        self.btn_plot.hide()
        self.btn_back.show()

        # אסוף את כל הנתונים מהטבלה
        points: list[tuple[datetime, float]] = []
        for r in range(self.table.rowCount()):
            date_item = self.table.item(r, 4)
            weight_item = self.table.item(r, 3)
            try:
                wval = float(weight_item.text().split()[0].replace(",", ".")) if weight_item else 0.0
                dval = datetime.strptime(date_item.text().strip(), "%d/%m/%Y") if date_item else datetime.now()
                points.append((dval, wval))
            except (ValueError, AttributeError):
                continue

        if not points:
            self._show_status("אין רשומות להצגה")
            return

        # מיין לפי תאריך
        points.sort(key=lambda x: x[0])
        xs = [p[0] for p in points]
        ys = [p[1] for p in points]

        # צייר גרף קווי עם ציר תאריכים
        self.figure.clear()
        # הגדר סגנון גרף
        self.figure.patch.set_facecolor('#ffffff')
        ax = self.figure.add_subplot(111)
        ax.set_facecolor('#f8f9fa')
        
        dates = mdates.date2num(xs)
        
        # ציור הקו הבסיסי
        ax.plot(dates, ys, '-', color='#2196F3', linewidth=3, alpha=0.7)
        
        # הוספת נקודות צבעוניות לפי עלייה/ירידה/ללא שינוי
        for i, y in enumerate(ys):
            if i == 0:
                color, size = '#2196F3', 120
            elif y > ys[i-1]:
                color, size = '#4CAF50', 140
            elif y < ys[i-1]:
                color, size = '#f44336', 140
            else:
                color, size = '#FF9800', 120
            
            ax.scatter(dates[i], y, s=size, c=color, marker='o', 
                      edgecolors='white', linewidths=2.5, zorder=5, alpha=0.9)
        
        ax.xaxis.set_major_locator(mdates.AutoDateLocator())
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m/%Y'))
        self.figure.autofmt_xdate(rotation=30)
        
        # כותרת מעוצבת וגדולה יותר - מסגרת קטנה יותר
        LRM = '\u200E'
        title = f"גרף משקלים - {self.exercise_name}"
        ax.set_title(f"{LRM}{title[::-1]}", 
                    fontsize=18,           # גודל גדול יותר
                    fontweight='bold',     # מודגש
                    pad=20,                # ריווח מהגרף
                    color='#1976D2',       # צבע כחול כהה
                    bbox=dict(boxstyle='round,pad=0.5', facecolor='#E3F2FD', 
                             edgecolor='#2196F3', linewidth=1.5))  # מסגרת קטנה יותר
        
        # הוספת kg למספרים על ציר Y
        from matplotlib.ticker import FuncFormatter
        def kg_formatter(x, pos):
            return f'{int(x)} kg'
        ax.yaxis.set_major_formatter(FuncFormatter(kg_formatter))
        
        # הגדרת רשת עדינה ויפה יותר
        ax.grid(True, linestyle='--', alpha=0.4, color='#BDBDBD', linewidth=0.8)
        ax.grid(True, which='minor', linestyle=':', alpha=0.2, color='#E0E0E0')
        ax.set_axisbelow(True)  # הרשת מאחורי הנתונים
        
        # עיצוב שולי הגרף - מסגרת מעוצבת יותר
        spine_colors = {
            'top': '#64B5F6',
            'bottom': '#1976D2', 
            'left': '#1976D2',
            'right': '#64B5F6'
        }
        
        for position, spine in ax.spines.items():
            spine.set_color(spine_colors.get(position, '#90A4AE'))
            spine.set_linewidth(2.5)
            spine.set_capstyle('round')
            
        # התאמת צבע וגודל תוויות הצירים
        ax.tick_params(axis='both', colors='#424242', labelsize=10, width=1.5, length=6)
        ax.tick_params(axis='x', rotation=0)  # תיקון זווית
        
        # הוספת צל עדין לאזור הגרף
        ax.set_facecolor('#FAFAFA')
        
        # שיפור המרווחים
        self.figure.tight_layout(pad=2.0)
        
        self.canvas.draw()
        self.canvas.show()

    def save_state(self):
        state = {"rows": [[self.table.item(r, c).text() if self.table.item(r, c) else "" 
                          for c in range(self.table.columnCount())] 
                         for r in range(self.table.rowCount())]}

        path = Path.cwd() / f"exercise_{self.profile_name}_{self.exercise_name}.json"
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(state, f, ensure_ascii=False, indent=2)
            self._has_unsaved_changes = False
            self._show_status(f"נשמר ל־{path}")
        except Exception as e:
            self._show_status(f"שגיאה בשמירה: {e}")

    def load_state(self):
        path = Path.cwd() / f"exercise_{self.profile_name}_{self.exercise_name}.json"
        if not path.exists():
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                state = json.load(f)
            self.table.setRowCount(0)
            for row_data in state.get("rows", []):
                r = self.table.rowCount()
                self.table.insertRow(r)
                for c, val in enumerate(row_data):
                    item = QTableWidgetItem(str(val))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
                    self.table.setItem(r, c, item)
            self.btn_pop.setEnabled(self.table.rowCount() > 0)
            self._update_summary()
            self._show_status(f"טען מצב מ־{path}")
        except Exception as e:
            self._show_status(f"שגיאה בטעינה: {e}")

    def _show_table_context_menu(self, pos):
        menu = QMenu()
        act_delete = menu.addAction("מחק שורות נבחרות")
        action = menu.exec(self.table.viewport().mapToGlobal(pos))
        if action == act_delete:
            self.delete_selected_rows()

    def delete_selected_rows(self):
        selected = sorted({idx.row() for idx in self.table.selectedIndexes()}, reverse=True)
        if selected:  # רק אם יש שורות נבחרות
            # שמירה למחסנית Undo לפני השינוי
            self._save_state_to_undo()
            self._has_unsaved_changes = True
            for r in selected:
                self.table.removeRow(r)
            self.btn_pop.setEnabled(self.table.rowCount() > 0)
            self._update_summary()

    def restore_normal_view(self):
        """החזרת התצוגה למצב רגיל"""
        self.input_container.show()
        self.table.show()
        self.btn_add.show()
        self.btn_pop.show()
        self.btn_delete_row.show()
        self.btn_duplicate_row.show()
        self.btn_plot.show()
        self.btn_back.hide()
        self.canvas.hide()

    def _update_delete_button(self):
        """עדכון מצב כפתור מחיקת שורה בהתאם לבחירה"""
        selected_rows = len({idx.row() for idx in self.table.selectedIndexes()})
        self.btn_delete_row.setEnabled(selected_rows == 1)
        self.btn_duplicate_row.setEnabled(selected_rows == 1)
    
    def delete_selected_row(self):
        """מחיקת השורה הנבחרת"""
        selected_rows = {idx.row() for idx in self.table.selectedIndexes()}
        if len(selected_rows) != 1:
            return
            
        self._save_state_to_undo()
        self.table.removeRow(selected_rows.pop())
        self._has_unsaved_changes = True
        self.btn_pop.setEnabled(self.table.rowCount() > 0)
        self._update_summary()
        self._show_status("השורה נמחקה.")
    
    def duplicate_selected_row(self):
        """שכפול השורה הנבחרת"""
        selected_rows = {idx.row() for idx in self.table.selectedIndexes()}
        if len(selected_rows) != 1:
            return
            
        row = selected_rows.pop()
        self._save_state_to_undo()
        
        # שכפול הנתונים מהשורה הנבחרת
        row_data = [self.table.item(row, col).text() if self.table.item(row, col) else "" 
                   for col in range(self.table.columnCount())]
        
        # הוספת שורה חדשה עם הנתונים המשוכפלים
        new_row = self.table.rowCount()
        self.table.insertRow(new_row)
        
        for col, value in enumerate(row_data):
            new_item = QTableWidgetItem(value)
            new_item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(new_row, col, new_item)
        
        self._has_unsaved_changes = True
        self.btn_pop.setEnabled(True)
        self._update_summary()
        self._show_status("השורה שוכפלה.")
    
    def _save_state_to_undo(self):
        """שמירת המצב הנוכחי למחסנית ה-Undo לפני ביצוע פעולה"""
        # אם אנחנו בתהליך שחזור, לא נשמור
        if self._is_restoring:
            return
        
        # שומר את המצב הנוכחי לפני השינוי
        state = self._get_current_table_state()
        # אם זה המצב הראשון, או שהמצב שונה מהמצב האחרון במחסנית
        if not self._undo_stack or state != self._undo_stack[-1]:
            self._undo_stack.append(state)
            # שמירה של מקסימום 5+1 מצבים (כולל המצב הנוכחי)
            if len(self._undo_stack) > self._max_undo + 1:
                self._undo_stack.pop(0)
        # כאשר נעשית פעולה חדשה, מנקים את מחסנית ה-Redo
        self._redo_stack.clear()
    
    def _get_current_table_state(self):
        """קבלת המצב הנוכחי של הטבלה"""
        return [[self.table.item(r, c).text() if self.table.item(r, c) else "" 
                for c in range(self.table.columnCount())] 
               for r in range(self.table.rowCount())]
    
    def _restore_table_state(self, state):
        """שחזור מצב הטבלה"""
        self._is_restoring = True  # מסמן שאנחנו בתהליך שחזור
        try:
            self.table.setRowCount(0)
            for row_data in state:
                r = self.table.rowCount()
                self.table.insertRow(r)
                for c, val in enumerate(row_data):
                    item = QTableWidgetItem(str(val))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
                    self.table.setItem(r, c, item)
            self.btn_pop.setEnabled(self.table.rowCount() > 0)
            self._update_summary()
        finally:
            self._is_restoring = False  # מסיים את תהליך השחזור
    
    def undo(self):
        """ביטול הפעולה האחרונה"""
        if len(self._undo_stack) < 1:
            self._show_status("אין מה לבטל")
            return
        
        # שמירת המצב הנוכחי ל-Redo (רק אם עדיין לא שמרנו אותו)
        current_state = self._get_current_table_state()
        if not self._redo_stack or current_state != self._redo_stack[-1]:
            self._redo_stack.append(current_state)
            if len(self._redo_stack) > self._max_undo:
                self._redo_stack.pop(0)
        
        # שחזור המצב הקודם
        previous_state = self._undo_stack.pop()
        self._restore_table_state(previous_state)
        self._has_unsaved_changes = True
        self._show_status("בוטל", 1000)
    
    def redo(self):
        """שחזור הפעולה שבוטלה"""
        if not self._redo_stack:
            self._show_status("אין מה לשחזר")
            return
        
        # שמירת המצב הנוכחי ל-Undo
        current_state = self._get_current_table_state()
        if not self._undo_stack or current_state != self._undo_stack[-1]:
            self._undo_stack.append(current_state)
            if len(self._undo_stack) > self._max_undo + 1:
                self._undo_stack.pop(0)
        
        # שחזור המצב מ-Redo
        state = self._redo_stack.pop()
        self._restore_table_state(state)
        self._has_unsaved_changes = True
        self._show_status("שוחזר", 1000)
        
    def _edit_date_cell(self, row: int, column: int):
        if column != 4:  # עמודת תאריך היא 4
            self.table.clearSelection()
            return
        item = self.table.item(row, column)
        if item is None:
            return
        
        # קריאת התאריך הנוכחי
        current = item.text() if item is not None else datetime.now().strftime("%d/%m/%Y")
        try:
            current_date = datetime.strptime(current, "%d/%m/%Y")
        except Exception:
            current_date = datetime.now()
        
        # יצירת דיאלוג עם לוח שנה
        dialog = QDialog(self)
        dialog.setWindowTitle("בחר תאריך")
        dialog.setModal(True)
        
        layout = QVBoxLayout()
        
        # יצירת לוח שנה
        calendar = QCalendarWidget()
        calendar.setGridVisible(True)
        
        # הגבלה: לא ניתן לבחור תאריך עתידי
        today = QDate.currentDate()
        calendar.setMaximumDate(today)
        
        calendar.setSelectedDate(QDate(current_date.year, current_date.month, current_date.day))
        
        # תווית להצגת התאריך הנבחר
        date_label = QLabel()
        date_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        date_label.setStyleSheet("font-size: 12pt; padding: 10px; background-color: #E3F2FD; border-radius: 4px;")
        
        def update_label():
            selected = calendar.selectedDate()
            date_label.setText(f"תאריך נבחר: {selected.toString('dd/MM/yyyy')}")
        
        update_label()
        calendar.selectionChanged.connect(update_label)
        
        # כפתורי אישור וביטול
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        
        layout.addWidget(calendar)
        layout.addWidget(date_label)
        layout.addWidget(button_box)
        dialog.setLayout(layout)
        
        # הצגת הדיאלוג
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # שמירה למחסנית Undo לפני השינוי
            self._save_state_to_undo()
            
            selected = calendar.selectedDate()
            new_date = selected.toString("dd/MM/yyyy")
            item.setText(new_date)
            item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
            self._has_unsaved_changes = True
            # עדכון רוחב העמודה כדי שיתאים לתוכן
            self.table._equalize_columns()
            # נקה בחירה ופוקוס
            self.table.clearSelection()
            self.table.clearFocus()
            self.table.setCurrentCell(-1, -1)
            try:
                self.save_state()
            except Exception:
                pass
        else:
            self.table.clearSelection()
            self.table.clearFocus()


class ImageCropDialog(QDialog):
    """דיאלוג לחיתוך אזור עגול מתמונה"""
    def __init__(self, image_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle("חיתוך תמונת פרופיל")
        self.setModal(True)
        
        # טעינת התמונה המקורית
        self.original_pixmap = QPixmap(image_path)
        
        # הגדרת גודל התצוגה (מקסימום 500x500)
        max_display_size = 500
        display_pixmap = self.original_pixmap.scaled(
            max_display_size, max_display_size,
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation
        )
        
        # יחס בין התצוגה למקור
        self.scale_factor = display_pixmap.width() / self.original_pixmap.width()
        
        # משתנים לחיתוך
        self.crop_diameter = min(display_pixmap.width(), display_pixmap.height()) // 2
        self.crop_x = (display_pixmap.width() - self.crop_diameter) // 2
        self.crop_y = (display_pixmap.height() - self.crop_diameter) // 2
        
        # Layout
        layout = QVBoxLayout()
        
        # כותרת
        title_label = QLabel("✂️ בחר את האזור לחיתוך")
        title_label.setStyleSheet("font-size: 14pt; font-weight: bold; color: #2196F3; padding: 10px;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # הוראות
        instructions = QLabel("גרור את המעגל למיקום הרצוי, השתמש בגלגלת לשינוי גודל")
        instructions.setStyleSheet("font-size: 10pt; color: #666; padding: 5px;")
        instructions.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(instructions)
        
        # תווית התמונה
        self.image_label = QLabel()
        self.image_label.setFixedSize(display_pixmap.width(), display_pixmap.height())
        self.image_label.setPixmap(display_pixmap)
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.image_label.setStyleSheet("border: 2px solid #2196F3;")
        
        # שמירת התמונה המקורית לתצוגה
        self.display_pixmap = display_pixmap
        
        # הוספה למרכז
        image_container = QHBoxLayout()
        image_container.addStretch()
        image_container.addWidget(self.image_label)
        image_container.addStretch()
        layout.addLayout(image_container)
        
        # כפתורים
        button_layout = QHBoxLayout()
        
        cancel_button = QPushButton("❌ ביטול")
        cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 10px 20px;
                font-size: 11pt;
                border-radius: 5px;
            }
            QPushButton:hover { background-color: #d32f2f; }
        """)
        cancel_button.clicked.connect(self.reject)
        
        crop_button = QPushButton("✂️ חתוך")
        crop_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 10px 20px;
                font-size: 11pt;
                border-radius: 5px;
            }
            QPushButton:hover { background-color: #45a049; }
        """)
        crop_button.clicked.connect(self.accept)
        
        button_layout.addWidget(cancel_button)
        button_layout.addStretch()
        button_layout.addWidget(crop_button)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        # משתנים לגרירה
        self.dragging = False
        self.last_pos = None
        
        # התקנת event filter
        self.image_label.installEventFilter(self)
        
        # שרטוט ראשוני
        self.update_display()
    
    def eventFilter(self, obj, event):
        """טיפול באירועי עכבר"""
        if obj == self.image_label:
            if event.type() == event.Type.MouseButtonPress:
                # בדוק אם לחצו בתוך המעגל
                pos = event.position()
                center_x = self.crop_x + self.crop_diameter // 2
                center_y = self.crop_y + self.crop_diameter // 2
                distance = ((pos.x() - center_x) ** 2 + (pos.y() - center_y) ** 2) ** 0.5
                
                if distance <= self.crop_diameter // 2:
                    self.dragging = True
                    self.last_pos = pos
                    return True
                    
            elif event.type() == event.Type.MouseMove:
                if self.dragging and self.last_pos:
                    pos = event.position()
                    dx = pos.x() - self.last_pos.x()
                    dy = pos.y() - self.last_pos.y()
                    
                    # עדכון מיקום המעגל
                    self.crop_x += dx
                    self.crop_y += dy
                    
                    # הגבלה לגבולות התמונה
                    self.crop_x = max(0, min(self.crop_x, self.image_label.width() - self.crop_diameter))
                    self.crop_y = max(0, min(self.crop_y, self.image_label.height() - self.crop_diameter))
                    
                    self.last_pos = pos
                    self.update_display()
                    return True
                    
            elif event.type() == event.Type.MouseButtonRelease:
                self.dragging = False
                self.last_pos = None
                return True
                
            elif event.type() == event.Type.Wheel:
                # שינוי גודל המעגל עם גלגלת העכבר
                delta = event.angleDelta().y()
                change = 10 if delta > 0 else -10
                
                new_diameter = self.crop_diameter + change
                min_size = 50
                max_size = min(self.image_label.width(), self.image_label.height())
                
                if min_size <= new_diameter <= max_size:
                    # שמור על המרכז
                    center_x = self.crop_x + self.crop_diameter // 2
                    center_y = self.crop_y + self.crop_diameter // 2
                    
                    self.crop_diameter = new_diameter
                    
                    # עדכון מיקום כך שהמרכז יישאר באותו מקום
                    self.crop_x = center_x - self.crop_diameter // 2
                    self.crop_y = center_y - self.crop_diameter // 2
                    
                    # הגבלה לגבולות
                    self.crop_x = max(0, min(self.crop_x, self.image_label.width() - self.crop_diameter))
                    self.crop_y = max(0, min(self.crop_y, self.image_label.height() - self.crop_diameter))
                    
                    self.update_display()
                return True
                
        return super().eventFilter(obj, event)
    
    def update_display(self):
        """עדכון התצוגה עם המעגל"""
        # שכפול התמונה המקורית
        display_pixmap = self.display_pixmap.copy()
        
        # שימוש ב-context manager כדי להבטיח סגירה נכונה
        painter = QPainter(display_pixmap)
        try:
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            
            # ציור רקע כהה למחוץ למעגל
            painter.setBrush(QColor(0, 0, 0, 150))
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawRect(display_pixmap.rect())
            
            # ציור המעגל (החלק הנבחר) - מחיקת הרקע הכהה בתוכו
            path = QPainterPath()
            path.addEllipse(self.crop_x, self.crop_y, self.crop_diameter, self.crop_diameter)
            
            painter.setCompositionMode(QPainter.CompositionMode.CompositionMode_DestinationOut)
            painter.fillPath(path, QColor(0, 0, 0, 150))
            
            painter.setCompositionMode(QPainter.CompositionMode.CompositionMode_SourceOver)
            
            # ציור גבול המעגל
            pen = QPen(QColor("#2196F3"), 3)
            painter.setPen(pen)
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawEllipse(self.crop_x, self.crop_y, self.crop_diameter, self.crop_diameter)
        finally:
            painter.end()
        
        # עכשיו אפשר להציב את הpixmap בבטחה
        self.image_label.setPixmap(display_pixmap)
    
    def get_cropped_pixmap(self):
        """קבלת התמונה החתוכה"""
        # חישוב קואורדינטות במקור
        original_x = int(self.crop_x / self.scale_factor)
        original_y = int(self.crop_y / self.scale_factor)
        original_diameter = int(self.crop_diameter / self.scale_factor)
        
        # חיתוך מהתמונה המקורית
        cropped = self.original_pixmap.copy(
            original_x, original_y, 
            original_diameter, original_diameter
        )
        
        # יצירת תמונה עגולה
        result = QPixmap(original_diameter, original_diameter)
        result.fill(Qt.GlobalColor.transparent)
        
        painter = QPainter(result)
        try:
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
            
            path = QPainterPath()
            path.addEllipse(0, 0, original_diameter, original_diameter)
            painter.setClipPath(path)
            painter.drawPixmap(0, 0, cropped)
        finally:
            painter.end()
        
        return result


def create_circular_pixmap(image_path, size):
    """יצירת תמונה עגולה מתמונה מלבנית"""
    # טעינת התמונה המקורית
    source = QPixmap(image_path)
    
    # יצירת pixmap חדש עם רקע שקוף
    target = QPixmap(size, size)
    target.fill(Qt.GlobalColor.transparent)
    
    # יצירת painter
    painter = QPainter(target)
    try:
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        
        # יצירת מסלול עגול
        path = QPainterPath()
        path.addEllipse(0, 0, size, size)
        
        # קביעת אזור החיתוך למעגל
        painter.setClipPath(path)
        
        # שרטוט התמונה (מותאמת לגודל)
        scaled_source = source.scaled(
            size, size,
            Qt.AspectRatioMode.KeepAspectRatioByExpanding,
            Qt.TransformationMode.SmoothTransformation
        )
        
        # מרכוז התמונה
        x = (size - scaled_source.width()) // 2
        y = (size - scaled_source.height()) // 2
        painter.drawPixmap(x, y, scaled_source)
    finally:
        painter.end()
    
    return target


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        # הגדרות חלון ראשי
        self.setWindowTitle(get_version_string())
        self.setMinimumSize(QSize(800, 600))
        self.showMaximized()  # פתיחה במסך מלא
        
        # הגדרת אייקון החלון
        self._set_window_icon()

        # יצירת סטטוס בר
        self.setStatusBar(QStatusBar())

        # יצירת מיכל מרכזי
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout()
        main_widget.setLayout(layout)

        # הגדרת טאבים
        self.tab_widget = QTabWidget()
        self.tab_widget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tab_widget.customContextMenuRequested.connect(self._show_tab_context_menu)
        layout.addWidget(self.tab_widget)

        # יצירת סרגל כלים
        toolbar = QToolBar()
        self.addToolBar(toolbar)
        
        # כפתור שמירה בסרגל כלים עם קיצור מקלדת
        save_action = QAction("שמור", self)
        save_action.setShortcuts([QKeySequence("Ctrl+S"), QKeySequence("Ctrl+ד")])  # תמיכה באנגלית ועברית
        save_action.triggered.connect(self._save_current_tab)
        toolbar.addAction(save_action)
        
        # הוספת spacer כדי לדחוף את תמונת הפרופיל לצד ימין
        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        toolbar.addWidget(spacer)
        
        # קונטיינר לתמונה ושם פרופיל
        profile_container = QWidget()
        profile_layout = QHBoxLayout(profile_container)
        profile_layout.setContentsMargins(5, 0, 5, 0)
        profile_layout.setSpacing(10)
        
        # שם הפרופיל
        self.profile_name_label = QLabel()
        self.profile_name_label.setStyleSheet("""
            QLabel {
                font-size: 13pt;
                font-weight: bold;
                color: #2196F3;
                padding: 5px;
            }
        """)
        self.profile_name_label.setCursor(Qt.CursorShape.PointingHandCursor)
        self.profile_name_label.mousePressEvent = lambda e: self._show_profile_dialog()
        self.profile_name_label.setToolTip("לחץ לצפייה בפרופיל (Ctrl+P)")
        profile_layout.addWidget(self.profile_name_label)
        
        # תמונת פרופיל בצד ימין של הToolbar
        self.profile_image_widget = QLabel()
        self.profile_image_widget.setFixedSize(56, 56)
        self.profile_image_widget.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.profile_image_widget.setScaledContents(False)
        self.profile_image_widget.setCursor(Qt.CursorShape.PointingHandCursor)
        self.profile_image_widget.mousePressEvent = lambda e: self._show_profile_dialog()
        self.profile_image_widget.setToolTip("לחץ לצפייה בפרופיל (Ctrl+P)")
        profile_layout.addWidget(self.profile_image_widget)
        
        toolbar.addWidget(profile_container)

        # תפריט קובץ
        file_menu = self.menuBar().addMenu("קובץ")
        
        # פעולת Undo
        undo_action = QAction("אחורה", self)
        undo_action.setShortcut(QKeySequence("Ctrl+Z"))
        undo_action.triggered.connect(self._undo_current_tab)
        file_menu.addAction(undo_action)
        self.addAction(undo_action)  # הוספה לחלון עצמו כדי שקיצור המקלדת יעבוד
        
        # קיצור נוסף בעברית ל-Undo
        undo_shortcut_he = QShortcut(QKeySequence("Ctrl+ז"), self)
        undo_shortcut_he.activated.connect(self._undo_current_tab)
        
        # פעולת Redo
        redo_action = QAction("קדימה", self)
        redo_action.setShortcut(QKeySequence("Ctrl+Y"))
        redo_action.triggered.connect(self._redo_current_tab)
        file_menu.addAction(redo_action)
        self.addAction(redo_action)  # הוספה לחלון עצמו כדי שקיצור המקלדת יעבוד
        
        # קיצורים נוספים בעברית ל-Redo
        redo_shortcut_he = QShortcut(QKeySequence("Ctrl+ט"), self)
        redo_shortcut_he.activated.connect(self._redo_current_tab)
        
        # קיצורי מקלדת לפרופיל (ללא כפתור בtoolbar)
        profile_shortcut = QShortcut(QKeySequence("Ctrl+P"), self)
        profile_shortcut.activated.connect(self._show_profile_dialog)
        profile_shortcut_he = QShortcut(QKeySequence("Ctrl+פ"), self)
        profile_shortcut_he.activated.connect(self._show_profile_dialog)
        
        file_menu.addSeparator()
        
        # פעולת שמירה בתפריט (משתמש באותו Action כמו הסרגל)
        file_menu.addAction(save_action)
        
        # פעולת שחזור
        restore_action = QAction("שחזר", self)
        restore_action.setShortcuts([QKeySequence("Ctrl+R"), QKeySequence("Ctrl+ר")])  # תמיכה באנגלית ועברית
        restore_action.triggered.connect(self._restore_current_tab)
        file_menu.addAction(restore_action)
        
        # פעולת ייצוא לאקסל
        export_action = QAction("ייצא לאקסל", self)
        export_action.setShortcuts([QKeySequence("Ctrl+E"), QKeySequence("Ctrl+ק")])  # תמיכה באנגלית ועברית
        export_action.triggered.connect(self._export_to_excel)
        file_menu.addAction(export_action)
        
        file_menu.addSeparator()
        
        # פעולת עזרה
        help_action = QAction("עזרה", self)
        help_action.triggered.connect(self._show_help)
        file_menu.addAction(help_action)

        # תפריט עריכה
        edit_menu = self.menuBar().addMenu("עריכה")
        
        # פעולת הוספת תרגיל
        add_exercise_action = QAction("הוסף תרגיל", self)
        add_exercise_action.setShortcuts([QKeySequence("Ctrl+N"), QKeySequence("Ctrl+מ")])  # תמיכה באנגלית ועברית
        add_exercise_action.triggered.connect(self._add_exercise)
        edit_menu.addAction(add_exercise_action)

        # פעולת ניקוי עמוד נוכחי
        clear_current_action = QAction("מחק עמוד", self)
        clear_current_action.triggered.connect(self._clear_current_tab)
        edit_menu.addAction(clear_current_action)

        # פעולת ניקוי נתונים בעמוד הנוכחי
        clear_data_action = QAction("נקה עמוד", self)
        clear_data_action.triggered.connect(self._clear_current_tab_data)
        edit_menu.addAction(clear_data_action)

        # פעולת ניקוי כל העמודים
        clear_all_action = QAction("מחק הכל", self)
        clear_all_action.triggered.connect(self._clear_all_tabs)
        edit_menu.addAction(clear_all_action)

        # שמירה בסגירה
        self._closing = False
        
        # טעינת פרטי פרופיל
        self.current_profile_name = None  # שם הפרופיל הנוכחי
        self._load_profile()
        
        # בדיקת פרופיל בהפעלה ראשונה - יבוצע אחרי שהחלון יוצג
        # טעינת התרגילים תתבצע בסוף _check_first_run או ישירות אם יש פרופיל
        QTimer.singleShot(100, self._check_first_run)
    
    def _check_first_run(self):
        """בדיקה אם זו הפעלה ראשונה ואין פרופיל"""
        # בדוק אם יש פרופילים קיימים
        profiles = self._get_all_profiles()
        
        if not profiles:
            # אין פרופילים - זו הפעלה ראשונה!
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setWindowTitle("ברוך הבא! 👋")
            msg.setText("🎉 זו ההפעלה הראשונה של האפליקציה!\n\nכדי להתחיל, עליך ליצור פרופיל אישי.")
            msg.setInformativeText("הפרופיל מאפשר לך:\n• לנהל מספר משתמשים באפליקציה\n• לעקוב אחר ההתקדמות האישית שלך\n• לשמור את הנתונים בנפרד")
            msg.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg.exec()
            
            # פתח מיד את דיאלוג יצירת פרופיל
            self._create_first_profile()
        
        # טען תרגילים (לאחר שהפרופיל נקבע)
        self._reload_exercises()
    
    def _create_first_profile(self):
        """יצירת פרופיל ראשון"""
        while True:
            # בקש שם לפרופיל
            name, ok = QInputDialog.getText(
                self, 
                "יצירת פרופיל ראשון",
                "הכנס שם לפרופיל שלך:",
                QLineEdit.EchoMode.Normal,
                "הפרופיל שלי"
            )
            
            if ok and name.strip():
                # יצירת פרופיל חדש
                self.current_profile_name = name.strip()
                empty_profile = {
                    "name": "",
                    "height": "",
                    "weight": "",
                    "age": "",
                    "gender": "",
                    "profile_image": ""
                }
                self._save_profile(empty_profile, self.current_profile_name)
                self.profile_data = empty_profile
                
                # הצע למלא פרטים נוספים
                reply = QMessageBox.question(
                    self,
                    "מילוי פרטים",
                    "האם ברצונך למלא את הפרטים האישיים שלך כעת?\n\n(ניתן למלא גם מאוחר יותר דרך כפתור הפרופיל)",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.Yes
                )
                
                if reply == QMessageBox.StandardButton.Yes:
                    self._show_profile_edit()
                
                # הכל בסדר, צא מהלולאה
                break
            elif ok:
                # שם ריק - בקש שוב
                QMessageBox.warning(self, "שגיאה", "נא להכניס שם לפרופיל")
            else:
                # המשתמש ביטל - חייב ליצור פרופיל!
                reply = QMessageBox.critical(
                    self,
                    "פרופיל נדרש",
                    "לא ניתן להשתמש באפליקציה ללא פרופיל.\n\nהאם ברצונך ליצור פרופיל עכשיו?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.Yes
                )
                
                if reply == QMessageBox.StandardButton.No:
                    # המשתמש לא רוצה ליצור פרופיל - סגור את האפליקציה
                    self.close()
                    break
    
    def _set_window_icon(self):
        """יצירת והגדרת אייקון מקצועי לחלון"""
        try:
            from PySide6.QtGui import QIcon, QBrush, QLinearGradient
            
            # יצירת פיקסמאפ בגודל גדול יותר לאיכות טובה
            size = 128
            pixmap = QPixmap(size, size)
            pixmap.fill(Qt.GlobalColor.transparent)  # רקע שקוף
            
            painter = QPainter(pixmap)
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
            
            # יצירת רקע עם גרדיאנט
            gradient = QLinearGradient(0, 0, size, size)
            gradient.setColorAt(0, QColor(33, 150, 243))    # כחול בהיר #2196F3
            gradient.setColorAt(1, QColor(25, 118, 210))    # כחול כהה #1976D2
            
            # ציור עיגול עם גרדיאנט
            painter.setBrush(QBrush(gradient))
            painter.setPen(QPen(QColor(21, 101, 192), 3))  # מסגרת כחולה כהה
            painter.drawEllipse(QRectF(2, 2, size-4, size-4))
            
            # ציור משקולת מסוגננת
            painter.setPen(QPen(QColor(255, 255, 255), 6, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap))
            
            # הבר האמצעי של המשקולת
            bar_y = size // 2
            bar_left = size * 0.3
            bar_right = size * 0.7
            painter.drawLine(int(bar_left), bar_y, int(bar_right), bar_y)
            
            # המשקולות משני הצדדים
            weight_size = size * 0.15
            
            # משקולת שמאלית
            painter.setBrush(QBrush(QColor(255, 255, 255)))
            painter.setPen(QPen(QColor(224, 224, 224), 2))
            left_rect = QRectF(bar_left - weight_size, bar_y - weight_size, weight_size * 2, weight_size * 2)
            painter.drawEllipse(left_rect)
            
            # משקולת ימנית
            right_rect = QRectF(bar_right - weight_size, bar_y - weight_size, weight_size * 2, weight_size * 2)
            painter.drawEllipse(right_rect)
            
            # הוספת נקודות דקורטיביות על המשקולות
            painter.setPen(QPen(QColor(33, 150, 243), 2))
            for rect in [left_rect, right_rect]:
                center_x = rect.center().x()
                center_y = rect.center().y()
                # שלוש נקודות במרכז כל משקולת
                painter.drawPoint(QPointF(center_x, center_y - 5))
                painter.drawPoint(QPointF(center_x, center_y))
                painter.drawPoint(QPointF(center_x, center_y + 5))
            
            painter.end()
            
            # הגדרת האייקון
            icon = QIcon(pixmap)
            self.setWindowIcon(icon)
            
            # שמירת האייקון לקובץ (אופציונלי - לשימוש עתידי)
            try:
                icon_path = Path.cwd() / "app_icon.png"
                pixmap.save(str(icon_path), "PNG")
            except Exception:
                pass
                
        except Exception:
            pass  # אם נכשל, פשוט לא יהיה אייקון

    def _load_profile(self):
        """טעינת פרטי הפרופיל מקובץ"""
        # אם current_profile_name לא מוגדר (התחלת התוכנית),
        # נטען את הפרופיל הפעיל האחרון
        load_from_active_file = False
        if not hasattr(self, 'current_profile_name'):
            load_from_active_file = True
            self.current_profile_name = None
        elif self.current_profile_name is None:
            load_from_active_file = True
            
        if load_from_active_file:
            active_profile_path = Path.cwd() / "active_profile.json"
            if active_profile_path.exists():
                try:
                    with open(active_profile_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                        self.current_profile_name = data.get("active_profile")
                except Exception:
                    pass
        
        # אם אין פרופיל פעיל, ננסה לטעון את הפרופיל הישן (user_profile.json)
        if not self.current_profile_name:
            old_profile_path = Path.cwd() / "user_profile.json"
            if old_profile_path.exists():
                self.current_profile_name = "פרופיל ראשי"
                # העברת הפרופיל הישן לפורמט החדש
                try:
                    with open(old_profile_path, "r", encoding="utf-8") as f:
                        old_data = json.load(f)
                        if old_data.get("name"):  # אם יש נתונים בפרופיל הישן
                            self._save_profile(old_data, "פרופיל ראשי")
                except Exception:
                    pass
        
        # איפוס נתוני הפרופיל - חשוב! כדי שלא יישארו ערכים מהפרופיל הקודם
        self.profile_data = {
            "name": "",
            "height": "",
            "weight": "",
            "age": "",
            "gender": "",
            "profile_image": ""
        }
        
        # טעינת נתוני הפרופיל הנוכחי מהקובץ שלו
        if self.current_profile_name:
            profile_path = Path.cwd() / f"profile_{self.current_profile_name}.json"
            if profile_path.exists():
                try:
                    with open(profile_path, "r", encoding="utf-8") as f:
                        loaded_data = json.load(f)
                        # עדכון רק השדות שקיימים בקובץ
                        self.profile_data.update(loaded_data)
                except Exception:
                    pass
            
            # עדכון שם הפרופיל בכותרת החלון
            self.setWindowTitle(f"{get_version_string()} - {self.current_profile_name}")
            
            # עדכון תמונת הפרופיל ב-toolbar
            self._update_profile_image_widget()
        else:
            self.setWindowTitle(get_version_string())
            self._update_profile_image_widget()

    def _update_profile_image_widget(self):
        """עדכון תמונת הפרופיל ושם הפרופיל ב-toolbar"""
        if not hasattr(self, 'profile_image_widget'):
            return
        
        # עדכון שם הפרופיל
        if hasattr(self, 'profile_name_label'):
            self.profile_name_label.setText(self.current_profile_name or "")
            
        profile_image_path = self.profile_data.get("profile_image", "")
        if profile_image_path and Path(profile_image_path).exists():
            try:
                # טעינת התמונה המקורית
                original = QPixmap(profile_image_path)
                if original.isNull():
                    self._set_default_profile_image()
                    return
                
                # יצירת פיקסמפ סופי בגודל 56x56 עם שקיפות
                final_pixmap = QPixmap(56, 56)
                final_pixmap.fill(Qt.GlobalColor.transparent)
                
                painter = QPainter(final_pixmap)
                try:
                    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
                    painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
                    
                    # יצירת path עגול לתמונה (50x50 במרכז)
                    path = QPainterPath()
                    path.addEllipse(3, 3, 50, 50)
                    painter.setClipPath(path)
                    
                    # רינדור התמונה בגודל המתאים
                    scaled = original.scaled(50, 50, Qt.AspectRatioMode.KeepAspectRatioByExpanding, 
                                            Qt.TransformationMode.SmoothTransformation)
                    x_offset = (50 - scaled.width()) // 2 + 3
                    y_offset = (50 - scaled.height()) // 2 + 3
                    painter.drawPixmap(x_offset, y_offset, scaled)
                    
                    # ביטול ה-clip לציור הבורדר
                    painter.setClipping(False)
                    
                    # ציור בורדר עגול
                    pen = QPen(QColor("#2196F3"), 3)
                    painter.setPen(pen)
                    painter.setBrush(Qt.BrushStyle.NoBrush)
                    painter.drawEllipse(2, 2, 52, 52)
                finally:
                    painter.end()
                
                self.profile_image_widget.setPixmap(final_pixmap)
            except Exception:
                self._set_default_profile_image()
        else:
            self._set_default_profile_image()
    
    def _set_default_profile_image(self):
        """הגדרת תמונת פרופיל ברירת מחדל"""
        if not hasattr(self, 'profile_image_widget'):
            return
        
        # יצירת פיקסמפ עגול עם האייקון ברירת מחדל
        pixmap = QPixmap(56, 56)
        pixmap.fill(Qt.GlobalColor.transparent)
        
        painter = QPainter(pixmap)
        try:
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
            
            # ציור רקע עגול
            painter.setBrush(QColor("#E3F2FD"))
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawEllipse(3, 3, 50, 50)
            
            # ציור האימוג'י
            painter.setPen(QColor("#2196F3"))
            font = QFont()
            font.setPointSize(20)
            painter.setFont(font)
            painter.drawText(3, 3, 50, 50, Qt.AlignmentFlag.AlignCenter, "👤")
            
            # ציור בורדר עגול
            pen = QPen(QColor("#2196F3"), 3)
            painter.setPen(pen)
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawEllipse(2, 2, 52, 52)
        finally:
            painter.end()
        
        self.profile_image_widget.setPixmap(pixmap)

    def _save_profile(self, profile_data, profile_name=None):
        """שמירת פרטי הפרופיל לקובץ"""
        if profile_name is None:
            profile_name = self.current_profile_name
        
        if not profile_name:
            QMessageBox.warning(self, "שגיאה", "לא נבחר פרופיל")
            return
            
        profile_path = Path.cwd() / f"profile_{profile_name}.json"
        try:
            with open(profile_path, "w", encoding="utf-8") as f:
                json.dump(profile_data, f, ensure_ascii=False, indent=2)
            self.profile_data = profile_data
            self.current_profile_name = profile_name
            
            # שמירת הפרופיל הפעיל
            active_profile_path = Path.cwd() / "active_profile.json"
            with open(active_profile_path, "w", encoding="utf-8") as f:
                json.dump({"active_profile": profile_name}, f, ensure_ascii=False, indent=2)
            
            self.setWindowTitle(f"{get_version_string()} - {profile_name}")
            
            # עדכון תמונת הפרופיל ב-toolbar
            self._update_profile_image_widget()
            
            self.statusBar().showMessage("פרטי הפרופיל נשמרו בהצלחה", 2000)
        except Exception as e:
            QMessageBox.warning(self, "שגיאה", f"שגיאה בשמירת הפרופיל: {e}")
    
    def _get_all_profiles(self):
        """קבלת רשימת כל הפרופילים"""
        profiles = []
        for file in Path.cwd().glob("profile_*.json"):
            profile_name = file.stem.replace("profile_", "")
            profiles.append(profile_name)
        return sorted(profiles)
    
    def _switch_profile(self):
        """החלפת פרופיל"""
        dialog = QDialog(self)
        dialog.setWindowTitle("החלף פרופיל")
        dialog.setModal(True)
        dialog.setMinimumWidth(400)
        
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        # תווית כותרת
        title_label = QLabel("🔄 בחר פרופיל או צור חדש")
        title_label.setStyleSheet("font-size: 16pt; font-weight: bold; color: #2196F3; padding: 10px;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # רשימת פרופילים קיימים
        profiles = self._get_all_profiles()
        
        if profiles:
            profiles_label = QLabel("פרופילים קיימים:")
            profiles_label.setStyleSheet("font-weight: bold; font-size: 12pt;")
            layout.addWidget(profiles_label)
            
            profiles_list = QListWidget()
            profiles_list.setStyleSheet("""
                QListWidget {
                    border: 2px solid #2196F3;
                    border-radius: 5px;
                    padding: 5px;
                    font-size: 11pt;
                }
                QListWidget::item {
                    padding: 8px;
                    border-radius: 3px;
                }
                QListWidget::item:selected {
                    background-color: #2196F3;
                    color: white;
                }
                QListWidget::item:hover {
                    background-color: #E3F2FD;
                }
            """)
            
            for profile in profiles:
                item = QListWidgetItem(f"👤 {profile}")
                if profile == self.current_profile_name:
                    item.setText(f"👤 {profile} (פעיל)")
                    item.setForeground(QColor("#4CAF50"))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                profiles_list.addItem(item)
            
            profiles_list.setMaximumHeight(200)
            layout.addWidget(profiles_list)
            
            # כפתור טעינת פרופיל
            load_button = QPushButton("✅ טען פרופיל נבחר")
            load_button.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    padding: 10px;
                    font-size: 12pt;
                    font-weight: bold;
                    border-radius: 5px;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)
            
            def load_selected_profile():
                current_item = profiles_list.currentItem()
                if current_item:
                    profile_text = current_item.text()
                    # הסרת האימוג'י וה"(פעיל)" אם קיים
                    profile_name = profile_text.replace("👤 ", "").replace(" (פעיל)", "").strip()
                    if profile_name != self.current_profile_name:
                        # בדיקה אם יש שינויים שלא נשמרו
                        has_unsaved = False
                        for i in range(self.tab_widget.count()):
                            tab = self.tab_widget.widget(i)
                            if isinstance(tab, ExerciseTab) and tab._has_unsaved_changes:
                                has_unsaved = True
                                break
                        
                        # אם יש שינויים, שאל את המשתמש
                        if has_unsaved:
                            reply = QMessageBox.question(
                                dialog,
                                "שינויים לא נשמרו",
                                f"⚠️ יש שינויים שלא נשמרו בפרופיל הנוכחי!\n\nהאם ברצונך לשמור לפני ההחלפה לפרופיל '{profile_name}'?",
                                QMessageBox.StandardButton.Save | QMessageBox.StandardButton.Discard | QMessageBox.StandardButton.Cancel,
                                QMessageBox.StandardButton.Save
                            )
                            
                            if reply == QMessageBox.StandardButton.Cancel:
                                return  # ביטול ההחלפה
                            elif reply == QMessageBox.StandardButton.Save:
                                # שמירת כל הטאבים עם שינויים
                                for i in range(self.tab_widget.count()):
                                    tab = self.tab_widget.widget(i)
                                    if isinstance(tab, ExerciseTab) and tab._has_unsaved_changes:
                                        try:
                                            tab.save_state()
                                        except Exception as e:
                                            QMessageBox.warning(dialog, "שגיאה בשמירה", f"שגיאה בשמירת {tab.exercise_name}: {e}")
                                            return
                        
                        # עדכון הפרופיל הנוכחי
                        self.current_profile_name = profile_name
                        
                        # שמירת הפרופיל הפעיל לקובץ
                        try:
                            active_profile_path = Path.cwd() / "active_profile.json"
                            with open(active_profile_path, "w", encoding="utf-8") as f:
                                json.dump({"active_profile": profile_name}, f, ensure_ascii=False, indent=2)
                        except Exception:
                            pass
                        
                        # טעינת נתוני הפרופיל
                        self._load_profile()
                        # טעינה מחדש של התרגילים
                        self._reload_exercises()
                        
                        QMessageBox.information(dialog, "הצלחה", f"הפרופיל '{profile_name}' נטען בהצלחה!")
                        dialog.accept()
                    else:
                        QMessageBox.information(dialog, "מידע", "פרופיל זה כבר פעיל")
                else:
                    QMessageBox.warning(dialog, "שגיאה", "נא לבחור פרופיל מהרשימה")
            
            load_button.clicked.connect(load_selected_profile)
            layout.addWidget(load_button)
            
            # כפתור מחיקת פרופיל
            delete_button = QPushButton("🗑️ מחק פרופיל נבחר")
            delete_button.setStyleSheet("""
                QPushButton {
                    background-color: #f44336;
                    color: white;
                    padding: 10px;
                    font-size: 12pt;
                    font-weight: bold;
                    border-radius: 5px;
                }
                QPushButton:hover {
                    background-color: #d32f2f;
                }
            """)
            
            def delete_selected_profile():
                current_item = profiles_list.currentItem()
                if not current_item:
                    QMessageBox.warning(dialog, "שגיאה", "נא לבחור פרופיל מהרשימה")
                    return
                
                profile_text = current_item.text()
                profile_name = profile_text.replace("👤 ", "").replace(" (פעיל)", "").strip()
                
                # אם זה הפרופיל הפעיל, לא ניתן למחוק
                if profile_name == self.current_profile_name:
                    QMessageBox.warning(dialog, "שגיאה", "לא ניתן למחוק את הפרופיל הפעיל הנוכחי.\nנא להחליף לפרופיל אחר לפני המחיקה.")
                    return
                
                # אישור מחיקה
                reply = QMessageBox.question(
                    dialog,
                    "אישור מחיקה",
                    f"האם אתה בטוח שברצונך למחוק את הפרופיל '{profile_name}'?\n\n⚠️ פעולה זו תמחק:\n• את פרטי הפרופיל\n• את כל נתוני התרגילים של הפרופיל\n\nהפעולה היא בלתי הפיכה!",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No
                )
                
                if reply == QMessageBox.StandardButton.Yes:
                    try:
                        # מחיקת קובץ הפרופיל
                        profile_path = Path.cwd() / f"profile_{profile_name}.json"
                        if profile_path.exists():
                            os.remove(profile_path)
                        
                        # מחיקת כל קבצי התרגילים של הפרופיל
                        for exercise_file in Path.cwd().glob(f"exercise_{profile_name}_*.json"):
                            try:
                                os.remove(exercise_file)
                            except Exception:
                                pass
                        
                        # הסרת הפרופיל מהרשימה
                        row = profiles_list.row(current_item)
                        profiles_list.takeItem(row)
                        
                        QMessageBox.information(dialog, "הצלחה", f"הפרופיל '{profile_name}' נמחק בהצלחה!")
                        
                        # אם אין יותר פרופילים, נסגור את הדיאלוג
                        if profiles_list.count() == 0:
                            QMessageBox.information(dialog, "מידע", "כל הפרופילים נמחקו.\nתוכל ליצור פרופיל חדש למטה.")
                    except Exception as e:
                        QMessageBox.warning(dialog, "שגיאה", f"שגיאה במחיקת הפרופיל: {e}")
            
            delete_button.clicked.connect(delete_selected_profile)
            layout.addWidget(delete_button)
            
            # כפתור עריכת שם פרופיל
            rename_button = QPushButton("✏️ ערוך שם פרופיל נבחר")
            rename_button.setStyleSheet("""
                QPushButton {
                    background-color: #FF9800;
                    color: white;
                    padding: 10px;
                    font-size: 12pt;
                    font-weight: bold;
                    border-radius: 5px;
                }
                QPushButton:hover {
                    background-color: #F57C00;
                }
            """)
            
            def rename_selected_profile():
                current_item = profiles_list.currentItem()
                if not current_item:
                    QMessageBox.warning(dialog, "שגיאה", "נא לבחור פרופיל מהרשימה")
                    return
                
                profile_text = current_item.text()
                old_name = profile_text.replace("👤 ", "").replace(" (פעיל)", "").strip()
                
                # בקש שם חדש
                new_name, ok = QInputDialog.getText(
                    dialog,
                    "עריכת שם פרופיל",
                    f"שם חדש עבור פרופיל '{old_name}':",
                    QLineEdit.EchoMode.Normal,
                    old_name
                )
                
                if ok and new_name.strip() and new_name != old_name:
                    # בדוק שאין פרופיל עם שם זהה
                    existing_profiles = self._get_all_profiles()
                    if new_name in existing_profiles:
                        QMessageBox.warning(dialog, "שגיאה", f"פרופיל בשם '{new_name}' כבר קיים!")
                        return
                    
                    try:
                        # שנה שם קובץ הפרופיל
                        old_profile_path = Path.cwd() / f"profile_{old_name}.json"
                        new_profile_path = Path.cwd() / f"profile_{new_name}.json"
                        
                        if old_profile_path.exists():
                            old_profile_path.rename(new_profile_path)
                        
                        # שנה שם כל קבצי התרגילים
                        for old_exercise_file in Path.cwd().glob(f"exercise_{old_name}_*.json"):
                            exercise_name = old_exercise_file.stem.replace(f"exercise_{old_name}_", "")
                            new_exercise_file = Path.cwd() / f"exercise_{new_name}_{exercise_name}.json"
                            old_exercise_file.rename(new_exercise_file)
                        
                        # אם זה הפרופיל הפעיל, עדכן את השם הפעיל
                        if old_name == self.current_profile_name:
                            self.current_profile_name = new_name
                            active_profile_path = Path.cwd() / "active_profile.json"
                            with open(active_profile_path, "w", encoding="utf-8") as f:
                                json.dump({"active_profile": new_name}, f, ensure_ascii=False, indent=2)
                            self.setWindowTitle(f"{get_version_string()} - {new_name}")
                        
                        # עדכן את הרשימה
                        item_text = f"👤 {new_name}"
                        if old_name == self.current_profile_name:
                            item_text += " (פעיל)"
                            current_item.setForeground(QColor("#4CAF50"))
                            font = current_item.font()
                            font.setBold(True)
                            current_item.setFont(font)
                        current_item.setText(item_text)
                        
                        QMessageBox.information(dialog, "הצלחה", f"שם הפרופיל שונה מ-'{old_name}' ל-'{new_name}'!")
                    except Exception as e:
                        QMessageBox.warning(dialog, "שגיאה", f"שגיאה בעריכת שם הפרופיל: {e}")
            
            rename_button.clicked.connect(rename_selected_profile)
            layout.addWidget(rename_button)
            
            # מפריד
            separator = QFrame()
            separator.setFrameShape(QFrame.Shape.HLine)
            separator.setStyleSheet("color: #ccc;")
            layout.addWidget(separator)
        
        # יצירת פרופיל חדש
        new_profile_label = QLabel("צור פרופיל חדש:")
        new_profile_label.setStyleSheet("font-weight: bold; font-size: 12pt;")
        layout.addWidget(new_profile_label)
        
        name_input = QLineEdit()
        name_input.setPlaceholderText("הכנס שם לפרופיל החדש")
        name_input.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                font-size: 11pt;
                border: 2px solid #2196F3;
                border-radius: 5px;
            }
        """)
        layout.addWidget(name_input)
        
        create_button = QPushButton("➕ צור פרופיל חדש")
        create_button.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 10px;
                font-size: 12pt;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        
        def create_new_profile():
            new_name = name_input.text().strip()
            if not new_name:
                QMessageBox.warning(dialog, "שגיאה", "נא להכניס שם לפרופיל")
                return
            
            if new_name in profiles:
                QMessageBox.warning(dialog, "שגיאה", "פרופיל בשם זה כבר קיים")
                return
            
            # בדיקה אם יש שינויים שלא נשמרו בפרופיל הנוכחי
            has_unsaved = False
            for i in range(self.tab_widget.count()):
                tab = self.tab_widget.widget(i)
                if isinstance(tab, ExerciseTab) and tab._has_unsaved_changes:
                    has_unsaved = True
                    break
            
            # אם יש שינויים, שאל את המשתמש
            if has_unsaved:
                reply = QMessageBox.question(
                    dialog,
                    "שינויים לא נשמרו",
                    f"⚠️ יש שינויים שלא נשמרו בפרופיל הנוכחי!\n\nהאם ברצונך לשמור לפני יצירת הפרופיל החדש '{new_name}'?",
                    QMessageBox.StandardButton.Save | QMessageBox.StandardButton.Discard | QMessageBox.StandardButton.Cancel,
                    QMessageBox.StandardButton.Save
                )
                
                if reply == QMessageBox.StandardButton.Cancel:
                    return  # ביטול יצירת הפרופיל
                elif reply == QMessageBox.StandardButton.Save:
                    # שמירת כל הטאבים עם שינויים
                    for i in range(self.tab_widget.count()):
                        tab = self.tab_widget.widget(i)
                        if isinstance(tab, ExerciseTab) and tab._has_unsaved_changes:
                            try:
                                tab.save_state()
                            except Exception as e:
                                QMessageBox.warning(dialog, "שגיאה בשמירה", f"שגיאה בשמירת {tab.exercise_name}: {e}")
                                return
            
            # יצירת פרופיל ריק חדש
            self.current_profile_name = new_name
            empty_profile = {
                "name": "",
                "height": "",
                "weight": "",
                "age": "",
                "gender": ""
            }
            self._save_profile(empty_profile, new_name)
            self.profile_data = empty_profile
            self._reload_exercises()  # טעינה מחדש של התרגילים (יהיה ריק)
            
            QMessageBox.information(dialog, "הצלחה", f"פרופיל '{new_name}' נוצר בהצלחה!\nכעת תוכל למלא את פרטי הפרופיל.")
            dialog.accept()
            # פתיחת חלון עריכת פרופיל
            self._show_profile_edit()
        
        create_button.clicked.connect(create_new_profile)
        layout.addWidget(create_button)
        
        # כפתור סגירה
        close_button = QPushButton("סגור")
        close_button.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                padding: 8px;
                font-size: 11pt;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        close_button.clicked.connect(dialog.reject)
        layout.addWidget(close_button)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def _reload_exercises(self):
        """טעינה מחדש של כל התרגילים לפרופיל הנוכחי"""
        # מחיקת כל הטאבים הקיימים
        while self.tab_widget.count() > 0:
            self.tab_widget.removeTab(0)
        
        # טעינת התרגילים של הפרופיל הנוכחי
        profile_name = self.current_profile_name or "ברירת מחדל"
        exercise_files = list(Path.cwd().glob(f"exercise_{profile_name}_*.json"))
        
        if exercise_files:
            for file in exercise_files:
                exercise_name = file.stem.replace(f"exercise_{profile_name}_", "")
                tab = ExerciseTab(exercise_name, profile_name)
                self.tab_widget.addTab(tab, exercise_name)
                tab.load_state()  # טעינת הנתונים
        else:
            # אם אין תרגילים, נציע ליצור אחד
            QMessageBox.information(self, "אין תרגילים", f"לפרופיל '{profile_name}' אין עדיין תרגילים.\nתוכל להוסיף תרגיל חדש דרך התפריט 'עריכה'.")
        
        # עדכן את גיליון הסיכום
        self._update_summary_tab()

    def _show_profile_dialog(self):
        """הצגת חלון עריכת פרופיל"""
        # בדיקה אם יש פרופיל קיים
        has_profile = any(self.profile_data.get(key, "") for key in ["name", "height", "weight", "age", "gender"])
        
        if has_profile:
            self._show_profile_view()
        else:
            self._show_profile_edit()
    
    def _show_profile_view(self):
        """הצגת פרופיל קיים במצב צפייה"""
        dialog = QDialog(self)
        dialog.setWindowTitle("פרופיל אישי")
        dialog.setModal(True)
        dialog.setMinimumWidth(450)
        
        layout = QVBoxLayout()
        layout.setSpacing(20)
        
        # תווית כותרת
        title_label = QLabel("📋 הפרופיל שלי")
        title_label.setStyleSheet("font-size: 18pt; font-weight: bold; color: #2196F3; padding: 15px;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # תמונת פרופיל
        profile_image_path = self.profile_data.get("profile_image", "")
        if profile_image_path and Path(profile_image_path).exists():
            image_label = QLabel()
            image_label.setFixedSize(120, 120)
            image_label.setStyleSheet("""
                QLabel {
                    border: 4px solid #2196F3;
                    border-radius: 60px;
                    background-color: #E3F2FD;
                }
            """)
            image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            image_label.setScaledContents(True)
            
            circular_pixmap = create_circular_pixmap(profile_image_path, 120)
            image_label.setPixmap(circular_pixmap)
            
            image_container = QHBoxLayout()
            image_container.addStretch()
            image_container.addWidget(image_label)
            image_container.addStretch()
            layout.addLayout(image_container)
        
        # תצוגת הפרטים
        info_widget = QWidget()
        info_widget.setStyleSheet("""
            QWidget {
                background-color: #f8f9fa;
                border-radius: 10px;
                padding: 20px;
            }
        """)
        info_layout = QVBoxLayout(info_widget)
        info_layout.setSpacing(15)
        
        # יצירת תוויות עם הפרטים
        profile_items = [
            ("👤 שם מלא:", self.profile_data.get("name", ""), "#E3F2FD", "#2196F3"),
            ("📏 גובה:", f"{self.profile_data.get('height', '')} ס\"מ" if self.profile_data.get('height') else "", "#E8F5E9", "#4CAF50"),
            ("⚖️ משקל:", f"{self.profile_data.get('weight', '')} ק\"ג" if self.profile_data.get('weight') else "", "#FFF3E0", "#FF9800"),
            ("🎂 גיל:", self.profile_data.get("age", ""), "#FCE4EC", "#E91E63"),
            ("👥 מין:", self.profile_data.get("gender", ""), "#F3E5F5", "#9C27B0")
        ]
        
        for label_text, value, bg_color, border_color in profile_items:
            if value:
                # יצירת מסגרת לכל פריט
                item_widget = QWidget()
                item_widget.setStyleSheet(f"""
                    QWidget {{
                        background-color: {bg_color};
                        border: 2px solid {border_color};
                        border-radius: 8px;
                        padding: 12px;
                    }}
                """)
                
                item_layout = QHBoxLayout(item_widget)
                item_layout.setContentsMargins(10, 8, 10, 8)
                
                label = QLabel(label_text)
                label.setStyleSheet(f"font-size: 13pt; font-weight: bold; color: {border_color}; border: none; background: transparent;")
                label.setMinimumWidth(120)
                
                value_label = QLabel(str(value))
                value_label.setStyleSheet("font-size: 13pt; font-weight: 600; color: #212529; border: none; background: transparent;")
                # יישור לימין רק לגיל
                if label_text.startswith("🎂"):
                    value_label.setAlignment(Qt.AlignmentFlag.AlignRight)
                else:
                    value_label.setAlignment(Qt.AlignmentFlag.AlignLeft)
                
                item_layout.addWidget(value_label)
                item_layout.addWidget(label)
                
                info_layout.addWidget(item_widget)
        
        layout.addWidget(info_widget)
        
        # כפתורים
        buttons_layout = QHBoxLayout()
        
        edit_button = QPushButton("✏️ ערוך פרופיל")
        edit_button.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 10px 20px;
                font-size: 12pt;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        edit_button.clicked.connect(lambda: (dialog.close(), self._show_profile_edit()))
        
        switch_button = QPushButton("🔄 החלף פרופיל")
        switch_button.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 10px 20px;
                font-size: 12pt;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
        """)
        switch_button.clicked.connect(lambda: (dialog.close(), self._switch_profile()))
        
        close_button = QPushButton("סגור")
        close_button.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                padding: 10px 20px;
                font-size: 12pt;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        close_button.clicked.connect(dialog.close)
        
        buttons_layout.addWidget(edit_button)
        buttons_layout.addWidget(switch_button)
        buttons_layout.addWidget(close_button)
        
        layout.addLayout(buttons_layout)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def _show_profile_edit(self):
        """הצגת טופס עריכת פרופיל"""
        dialog = QDialog(self)
        dialog.setWindowTitle("עריכת פרופיל אישי")
        dialog.setModal(True)
        dialog.setMinimumWidth(400)
        
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        # תווית כותרת
        title_label = QLabel("📋 פרטים אישיים")
        title_label.setStyleSheet("font-size: 16pt; font-weight: bold; color: #2196F3; padding: 10px;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # תמונת פרופיל
        image_layout = QHBoxLayout()
        
        # תצוגת התמונה
        profile_image_label = QLabel()
        profile_image_label.setFixedSize(100, 100)
        profile_image_label.setStyleSheet("""
            QLabel {
                border: 3px solid #2196F3;
                border-radius: 50px;
                background-color: #E3F2FD;
            }
        """)
        profile_image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        profile_image_label.setScaledContents(True)
        
        # טעינת תמונה קיימת או הצגת טקסט ברירת מחדל
        current_image_path = self.profile_data.get("profile_image", "")
        if current_image_path and Path(current_image_path).exists():
            circular_pixmap = create_circular_pixmap(current_image_path, 100)
            profile_image_label.setPixmap(circular_pixmap)
        else:
            profile_image_label.setText("📷\nאין תמונה")
            profile_image_label.setStyleSheet("""
                QLabel {
                    border: 3px dashed #2196F3;
                    border-radius: 50px;
                    background-color: #E3F2FD;
                    color: #2196F3;
                    font-size: 10pt;
                }
            """)
        
        image_layout.addStretch()
        image_layout.addWidget(profile_image_label)
        
        # כפתורי תמונה
        image_buttons_layout = QVBoxLayout()
        
        upload_image_button = QPushButton("📤 העלה תמונה")
        upload_image_button.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                font-size: 10pt;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        
        remove_image_button = QPushButton("🗑️ הסר תמונה")
        remove_image_button.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 8px;
                font-size: 10pt;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        """)
        
        selected_image_path = [current_image_path]  # משתמשים ברשימה כדי לעדכן מתוך פונקציה פנימית
        
        def upload_image():
            file_path, _ = QFileDialog.getOpenFileName(
                dialog,
                "בחר תמונת פרופיל",
                str(Path.home()),
                "תמונות (*.png *.jpg *.jpeg *.bmp *.gif)"
            )
            
            if file_path:
                try:
                    # פתיחת דיאלוג חיתוך
                    crop_dialog = ImageCropDialog(file_path, dialog)
                    if crop_dialog.exec() == QDialog.DialogCode.Accepted:
                        # קבלת התמונה החתוכה
                        cropped_pixmap = crop_dialog.get_cropped_pixmap()
                        
                        # שמירת התמונה החתוכה - תמיד כ-PNG (תומך בשקיפות)
                        new_image_path = Path.cwd() / f"profile_image_{self.current_profile_name}.png"
                        
                        # שמירה
                        if not cropped_pixmap.save(str(new_image_path), "PNG"):
                            raise Exception("שגיאה בשמירת הקובץ")
                        
                        selected_image_path[0] = str(new_image_path)
                        
                        # עדכון התצוגה
                        circular_pixmap = create_circular_pixmap(str(new_image_path), 100)
                        profile_image_label.setPixmap(circular_pixmap)
                        profile_image_label.setText("")
                        profile_image_label.setStyleSheet("""
                            QLabel {
                                border: 3px solid #2196F3;
                                border-radius: 50px;
                                background-color: #E3F2FD;
                            }
                        """)
                except Exception as e:
                    QMessageBox.warning(dialog, "שגיאה", f"שגיאה בהעלאת התמונה: {e}")
        
        def remove_image():
            selected_image_path[0] = ""
            profile_image_label.clear()
            profile_image_label.setText("📷\nאין תמונה")
            profile_image_label.setStyleSheet("""
                QLabel {
                    border: 3px dashed #2196F3;
                    border-radius: 50px;
                    background-color: #E3F2FD;
                    color: #2196F3;
                    font-size: 10pt;
                }
            """)
            profile_image_label.setScaledContents(False)
        
        upload_image_button.clicked.connect(upload_image)
        remove_image_button.clicked.connect(remove_image)
        
        image_buttons_layout.addWidget(upload_image_button)
        image_buttons_layout.addWidget(remove_image_button)
        image_buttons_layout.addStretch()
        
        image_layout.addLayout(image_buttons_layout)
        image_layout.addStretch()
        
        layout.addLayout(image_layout)
        
        # מפריד
        separator = QFrame()
        separator.setFrameShape(QFrame.Shape.HLine)
        separator.setStyleSheet("color: #ccc;")
        layout.addWidget(separator)
        
        # טופס הפרטים
        form_layout = QGridLayout()
        form_layout.setSpacing(10)
        
        # שם
        name_label = QLabel("שם מלא:")
        name_label.setStyleSheet("font-weight: bold;")
        name_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        name_input = QLineEdit()
        name_input.setPlaceholderText("הכנס את שמך המלא")
        name_input.setText(self.profile_data.get("name", ""))
        name_input.setAlignment(Qt.AlignmentFlag.AlignLeft)
        form_layout.addWidget(name_input, 0, 0)
        form_layout.addWidget(name_label, 0, 1)
        
        # גובה
        height_label = QLabel("גובה (ס\"מ):")
        height_label.setStyleSheet("font-weight: bold;")
        height_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        height_input = QLineEdit()
        height_input.setPlaceholderText("לדוגמה: 175")
        height_input.setValidator(QIntValidator(100, 250, self))
        height_input.setText(self.profile_data.get("height", ""))
        height_input.setAlignment(Qt.AlignmentFlag.AlignRight)
        form_layout.addWidget(height_input, 1, 0)
        form_layout.addWidget(height_label, 1, 1)
        
        # משקל
        weight_label = QLabel("משקל (ק\"ג):")
        weight_label.setStyleSheet("font-weight: bold;")
        weight_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        weight_input = QLineEdit()
        weight_input.setPlaceholderText("לדוגמה: 75.5")
        weight_input.setValidator(QDoubleValidator(30.0, 300.0, 1, self))
        weight_input.setText(self.profile_data.get("weight", ""))
        weight_input.setAlignment(Qt.AlignmentFlag.AlignRight)
        form_layout.addWidget(weight_input, 2, 0)
        form_layout.addWidget(weight_label, 2, 1)
        
        # גיל
        age_label = QLabel("גיל:")
        age_label.setStyleSheet("font-weight: bold;")
        age_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        age_input = QLineEdit()
        age_input.setPlaceholderText("לדוגמה: 25")
        age_input.setValidator(QIntValidator(10, 120, self))
        age_input.setText(self.profile_data.get("age", ""))
        age_input.setAlignment(Qt.AlignmentFlag.AlignRight)
        form_layout.addWidget(age_input, 3, 0)
        form_layout.addWidget(age_label, 3, 1)
        
        # מין
        gender_label = QLabel("מין:")
        gender_label.setStyleSheet("font-weight: bold;")
        gender_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        gender_layout = QHBoxLayout()
        
        gender_group = QButtonGroup(dialog)
        male_radio = QRadioButton("זכר")
        female_radio = QRadioButton("נקבה")
        gender_group.addButton(male_radio)
        gender_group.addButton(female_radio)
        
        current_gender = self.profile_data.get("gender", "")
        if current_gender == "זכר":
            male_radio.setChecked(True)
        elif current_gender == "נקבה":
            female_radio.setChecked(True)
        
        gender_layout.addStretch()
        gender_layout.addWidget(female_radio)
        gender_layout.addWidget(male_radio)
        
        form_layout.addLayout(gender_layout, 4, 0)
        form_layout.addWidget(gender_label, 4, 1)
        
        layout.addLayout(form_layout)
        
        # כפתורי פעולה
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        button_box.button(QDialogButtonBox.StandardButton.Save).setText("שמור")
        button_box.button(QDialogButtonBox.StandardButton.Cancel).setText("ביטול")
        
        def save_profile():
            # בדיקת תקינות
            if not name_input.text().strip():
                QMessageBox.warning(dialog, "שגיאה", "נא למלא שם")
                return
            
            # קבלת המין שנבחר
            selected_gender = ""
            if male_radio.isChecked():
                selected_gender = "זכר"
            elif female_radio.isChecked():
                selected_gender = "נקבה"
            
            profile_data = {
                "name": name_input.text().strip(),
                "height": height_input.text().strip(),
                "weight": weight_input.text().strip(),
                "age": age_input.text().strip(),
                "gender": selected_gender,
                "profile_image": selected_image_path[0]
            }
            self._save_profile(profile_data)
            dialog.accept()
            # הצגת מסך הפרופיל אחרי השמירה
            self._show_profile_view()
        
        button_box.accepted.connect(save_profile)
        button_box.rejected.connect(dialog.reject)
        
        layout.addWidget(button_box)
        
        dialog.setLayout(layout)
        dialog.exec()

    def _add_exercise(self):
        title, ok = QInputDialog.getText(self, "הוספת תרגיל", "שם התרגיל:")
        if ok and title.strip():
            # בדוק אם תרגיל עם שם זהה כבר קיים
            existing = set()
            for i in range(self.tab_widget.count()):
                tab = self.tab_widget.widget(i)
                if isinstance(tab, ExerciseTab):
                    existing.add(tab.exercise_name)
            if title not in existing:
                tab = ExerciseTab(title, self.current_profile_name)
                self.tab_widget.addTab(tab, title)
                self.tab_widget.setCurrentWidget(tab)
                # עדכן את גיליון הסיכום
                self._update_summary_tab()

    def _show_tab_context_menu(self, position):
        """הצגת תפריט הקשר על טאב - רק לתרגילים"""
        # מציאת הטאב שנלחץ עליו
        tab_bar = self.tab_widget.tabBar()
        index = tab_bar.tabAt(position)
        
        if index == -1:
            return
        
        # בדוק שזה ExerciseTab ולא SummaryTab
        tab = self.tab_widget.widget(index)
        if not isinstance(tab, ExerciseTab):
            return  # אל תציג תפריט לגיליון סיכום
        
        # יצירת תפריט
        menu = QMenu(self)
        
        # פעולת שינוי שם
        rename_action = QAction("✏️ שנה שם", self)
        rename_action.triggered.connect(lambda: self._rename_exercise(index))
        menu.addAction(rename_action)
        
        # הצגת התפריט במיקום העכבר
        menu.exec(tab_bar.mapToGlobal(position))
    
    def _rename_exercise(self, index):
        """שינוי שם תרגיל"""
        tab = self.tab_widget.widget(index)
        if not isinstance(tab, ExerciseTab):
            return
        
        old_name = tab.exercise_name
        
        # בקש שם חדש
        new_name, ok = QInputDialog.getText(
            self,
            "שינוי שם תרגיל",
            f"שם חדש עבור '{old_name}':",
            QLineEdit.EchoMode.Normal,
            old_name
        )
        
        if ok and new_name.strip() and new_name != old_name:
            # בדוק שאין תרגיל עם שם זהה
            existing = set()
            for i in range(self.tab_widget.count()):
                t = self.tab_widget.widget(i)
                if isinstance(t, ExerciseTab) and i != index:
                    existing.add(t.exercise_name)
            
            if new_name in existing:
                QMessageBox.warning(self, "שגיאה", f"תרגיל בשם '{new_name}' כבר קיים!")
                return
            
            # שנה את שם הקובץ
            old_path = Path.cwd() / f"exercise_{self.current_profile_name}_{old_name}.json"
            new_path = Path.cwd() / f"exercise_{self.current_profile_name}_{new_name}.json"
            
            try:
                if old_path.exists():
                    old_path.rename(new_path)
                
                # עדכן את הטאב
                tab.exercise_name = new_name
                self.tab_widget.setTabText(index, new_name)
                
                self.statusBar().showMessage(f"שם התרגיל שונה ל-'{new_name}'", 2000)
            except Exception as e:
                QMessageBox.warning(self, "שגיאה", f"שגיאה בשינוי שם: {e}")

    def _update_summary_tab(self):
        """עדכון גיליון הסיכום - מוצג רק אם יש לפחות 2 תרגילים"""
        # ספור תרגילים (לא כולל גיליון סיכום אם קיים)
        exercise_count = 0
        summary_tab_index = -1
        
        for i in range(self.tab_widget.count()):
            tab = self.tab_widget.widget(i)
            if isinstance(tab, SummaryTab):
                summary_tab_index = i
            elif isinstance(tab, ExerciseTab):
                exercise_count += 1
        
        # אם יש 2 תרגילים או יותר ואין גיליון סיכום - צור אותו
        if exercise_count >= 2 and summary_tab_index == -1:
            summary_tab = SummaryTab()
            self.tab_widget.insertTab(0, summary_tab, "📊 סיכום")
        
        # אם יש פחות מ-2 תרגילים וקיים גיליון סיכום - הסר אותו
        elif exercise_count < 2 and summary_tab_index != -1:
            self.tab_widget.removeTab(summary_tab_index)

    def _save_current_tab(self):
        current = self.tab_widget.currentWidget()
        if isinstance(current, ExerciseTab):
            try:
                current.save_state()
            except Exception as e:
                QMessageBox.warning(self, "שגיאה בשמירה", str(e))

    def _restore_current_tab(self):
        current = self.tab_widget.currentWidget()
        if isinstance(current, ExerciseTab):
            try:
                current.load_state()
                self.statusBar().showMessage("שוחזר בהצלחה מקובץ", 2000)
            except Exception as e:
                QMessageBox.warning(self, "שגיאה בשחזור", str(e))
    
    def _export_to_excel(self):
        """ייצוא כל העמודים לקובץ אקסל, כל עמוד לגיליון נפרד"""
        # בדוק אם openpyxl מותקן
        if not _HAS_OPENPYXL:
            QMessageBox.critical(self, "שגיאה", "openpyxl לא מותקן.\n\nכדי לייצא לאקסל, התקן את החבילה:\npip install openpyxl")
            return
        
        # בדוק אם יש עמודים לייצא
        if self.tab_widget.count() == 0:
            QMessageBox.warning(self, "שגיאה", "אין עמודים לייצוא")
            return
        
        # צור שם קובץ ברירת מחדל
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"תרגילים_{timestamp}.xlsx"
        
        # בקש מהמשתמש שם קובץ
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "שמור קובץ אקסל",
            default_filename,
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return  # המשתמש ביטל
        
        try:
            # צור חוברת עבודה חדשה
            wb = Workbook()
            # הסר את הגיליון הראשון שנוצר אוטומטית
            if wb.active:
                wb.remove(wb.active)
            
            # עבור על כל העמודים באפליקציה
            for tab_index in range(self.tab_widget.count()):
                tab = self.tab_widget.widget(tab_index)
                if not isinstance(tab, ExerciseTab):
                    continue
                
                exercise_name = tab.exercise_name
                
                # צור גיליון חדש לתרגיל הזה
                ws = wb.create_sheet(title=exercise_name[:31])  # שם גיליון מוגבל ל-31 תווים
                
                # הגדר את הגיליון להיות מימין לשמאל (RTL)
                ws.sheet_view.rightToLeft = True
                
                # הוסף כותרות
                headers = ["תאריך", "משקל", "סטים", "חזרות", "סט אחרון"]
                ws.append(headers)
                
                # עצב את שורת הכותרת
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # הוסף נתונים מהטבלה
                table = tab.table
                for row in range(table.rowCount()):
                    row_data = []
                    
                    # קרא את כל הערכים מהשורה
                    values = []
                    for col in range(table.columnCount()):
                        item = table.item(row, col)
                        values.append(item.text() if item else "")
                    
                    # הוסף בסדר הפוך: תאריך (4), משקל (3), סטים (2), חזרות (1), סט אחרון (0)
                    for col_index in [4, 3, 2, 1, 0]:
                        text = values[col_index]
                        if not text:
                            row_data.append("")
                            continue
                        
                        # עמודה 4 מהטבלה היא תאריך
                        if col_index == 4:
                            try:
                                date_obj = datetime.strptime(text, "%d/%m/%Y")
                                row_data.append(date_obj)
                            except ValueError:
                                row_data.append(text)
                        # עמודה 3 היא משקל - נסיר את "kg"
                        elif col_index == 3:
                            clean_text = text.replace("kg", "").replace("KG", "").strip()
                            parts = text.split()
                            if parts:
                                clean_text = parts[0].replace(",", ".")
                            
                            try:
                                if '.' in clean_text or ',' in clean_text:
                                    row_data.append(float(clean_text.replace(",", ".")))
                                else:
                                    row_data.append(int(clean_text))
                            except (ValueError, AttributeError):
                                try:
                                    import re
                                    number_str = re.search(r'[\d.,]+', text)
                                    if number_str:
                                        num_text = number_str.group().replace(",", ".")
                                        if '.' in num_text:
                                            row_data.append(float(num_text))
                                        else:
                                            row_data.append(int(num_text))
                                    else:
                                        row_data.append(0)
                                except Exception:
                                    row_data.append(0)
                        # עמודות 0-2 הן מספרים אחרים
                        else:
                            try:
                                if '.' in text:
                                    row_data.append(float(text))
                                else:
                                    row_data.append(int(text))
                            except ValueError:
                                row_data.append(text)
                    
                    ws.append(row_data)
                
                # הפוך את הטבלה לטבלה חכמה של Excel
                if _HAS_OPENPYXL:
                    try:
                        from openpyxl.worksheet.table import Table, TableStyleInfo
                    except ImportError:
                        pass
                
                max_row = ws.max_row
                max_col = ws.max_column
                if max_row > 1 and _HAS_OPENPYXL:
                    try:
                        table_range = f"A1:{get_column_letter(max_col)}{max_row}"
                        excel_table = Table(displayName=f"DataTable{tab_index}", ref=table_range)
                        
                        style = TableStyleInfo(
                            name="TableStyleMedium2",
                            showFirstColumn=False,
                            showLastColumn=False,
                            showRowStripes=True,
                            showColumnStripes=False
                        )
                        excel_table.tableStyleInfo = style
                        ws.add_table(excel_table)
                    except Exception:
                        pass  # אם הטבלה נכשלת, נמשיך בלי
                
                # התאם רוחב עמודות
                for col in range(1, max_col + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 15
                
                # עצב את עמודת התאריך
                for row in range(2, max_row + 1):
                    date_cell = ws.cell(row=row, column=1)
                    if date_cell.value and isinstance(date_cell.value, datetime):
                        date_cell.number_format = 'DD/MM/YYYY'
                
                # צור גרף קווי
                if max_row > 1 and _HAS_OPENPYXL:
                    try:
                        chart = LineChart()
                        chart.title = f"גרף משקלים - {exercise_name}"
                        chart.style = 10
                        chart.y_axis.title = None
                        chart.x_axis.title = None
                        chart.legend = None
                        
                        data = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=max_row)
                        dates = Reference(ws, min_col=1, min_row=2, max_row=max_row)
                        
                        chart.add_data(data, titles_from_data=False)
                        chart.set_categories(dates)
                        
                        if len(chart.series) > 0:
                            series = chart.series[0]
                            series.smooth = False
                            
                            try:
                                from openpyxl.chart.marker import Marker
                                from openpyxl.drawing.line import LineProperties
                                
                                line = LineProperties()
                                line.solidFill = "2196F3"
                                line.width = 25000
                                series.graphicalProperties.line = line
                                
                                marker = Marker(symbol='circle', size=5)
                                series.marker = marker
                            except ImportError:
                                pass  # אם לא ניתן לעצב, נמשיך בלי
                        
                        chart.width = 20
                        chart.height = 12
                        
                        ws.add_chart(chart, f"A{max_row + 3}")
                    except Exception:
                        pass  # אם הגרף נכשל, נמשיך בלי
            
            # שמור את הקובץ
            wb.save(filename)
            
            self.statusBar().showMessage(f"נשמר בהצלחה: {filename}", 3000)
            QMessageBox.information(self, "הצלחה", f"הקובץ נשמר בהצלחה:\n{filename}\n\nיוצאו {self.tab_widget.count()} תרגילים")
            
        except Exception as e:
            QMessageBox.critical(self, "שגיאה", f"שגיאה בשמירת הקובץ:\n{str(e)}")
    
    def _undo_current_tab(self):
        """ביטול הפעולה האחרונה בעמוד הנוכחי"""
        current = self.tab_widget.currentWidget()
        if isinstance(current, ExerciseTab):
            current.undo()
    
    def _redo_current_tab(self):
        """שחזור הפעולה שבוטלה בעמוד הנוכחי"""
        current = self.tab_widget.currentWidget()
        if isinstance(current, ExerciseTab):
            current.redo()

    def _show_help(self):
        """הצגת חלון עזרה עם מידע על האפליקציה"""
        help_text = """
        <div dir="rtl" style="text-align: left; font-size: 11pt; direction: rtl;">
        <h2 style="text-align: left;">אפליקציית מעקב משקלים</h2>
        <p style="text-align: left;">אפליקציה לניהול ומעקב אחר התקדמות באימוני כוח.</p>
        
        <h3 style="text-align: left;">תכונות עיקריות:</h3>
        <ul style="text-align: left;">
            <li style="text-align: left;"><b>ניהול תרגילים מרובים</b> - ניתן ליצור עמודים נפרדים לכל תרגיל</li>
            <li style="text-align: left;"><b>מעקב מפורט</b> - רישום משקל, מספר סטים, חזרות וסט אחרון</li>
            <li style="text-align: left;"><b>גרפים ויזואליים</b> - הצגת התקדמות לאורך זמן</li>
            <li style="text-align: left;"><b>חישוב סטטיסטיקות</b> - סיכום אימונים וסך משקל מצטבר</li>
            <li style="text-align: left;"><b>שמירה אוטומטית</b> - כל הנתונים נשמרים למחשב</li>
        </ul>
        
        <h3 style="text-align: left;">קיצורי מקלדת:</h3>
        <ul style="text-align: left;">
            <li style="text-align: left;"><b>Ctrl+Z</b> - אחורה (ביטול פעולה)</li>
            <li style="text-align: left;"><b>Ctrl+Y</b> - קדימה (שחזור פעולה)</li>
            <li style="text-align: left;"><b>Ctrl+S</b> - שמור</li>
            <li style="text-align: left;"><b>Ctrl+R</b> - שחזר מקובץ</li>
            <li style="text-align: left;"><b>Ctrl+N</b> - הוסף תרגיל חדש</li>
            <li style="text-align: left;"><b>Enter</b> - הוסף רשומה (כשכל השדות מלאים)</li>
            <li style="text-align: left;"><b>חיצים ↑↓</b> - מעבר בין שדות קלט</li>
        </ul>
        
        <h3 style="text-align: left;">טיפים:</h3>
        <ul style="text-align: left;">
            <li style="text-align: left;">לחץ פעמיים על תאריך לעריכה</li>
            <li style="text-align: left;">בחר שורה ולחץ "מחק שורה" למחיקה</li>
            <li style="text-align: left;">השתמש ב"הצג גרף" לראות התקדמות ויזואלית</li>
        </ul>
        
        <p style="margin-top: 20px; color: #666; text-align: left;">
        גרסה 1.0 | 2025
        </p>
        </div>
        """
        
        msg = QMessageBox(self)
        msg.setWindowTitle("עזרה - מעקב משקלים")
        msg.setTextFormat(Qt.TextFormat.RichText)
        msg.setText(help_text)
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

    def _clear_current_tab_data(self):
        """ניקוי כל הנתונים מהעמוד הנוכחי אבל שמירת העמוד עצמו"""
        current = self.tab_widget.currentWidget()
        if not isinstance(current, ExerciseTab):
            return
            
        reply = QMessageBox.question(
            self,
            "אישור ניקוי נתונים",
            f"האם אתה בטוח שברצונך למחוק את כל הנתונים מהעמוד '{current.exercise_name}'?\n\nהעמוד יישאר קיים אך ללא נתונים.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                # מחיקת כל השורות מהטבלה
                current.table.setRowCount(0)
                
                # איפוס כפתורי המחיקה
                current.btn_pop.setEnabled(False)
                current.btn_delete_row.setEnabled(False)
                
                # עדכון הסיכום
                current._update_summary()
                
                # סימון שיש שינויים לא שמורים
                current._has_unsaved_changes = True
                
                # מחיקת קובץ השמירה
                path = Path.cwd() / f"exercise_state_{current.exercise_name}.json"
                if path.exists():
                    os.remove(path)
                
                self.statusBar().showMessage(f"נמחקו כל הנתונים מהעמוד '{current.exercise_name}'", 2000)
            except Exception as e:
                QMessageBox.warning(self, "שגיאה בניקוי", str(e))

    def _clear_current_tab(self):
        current = self.tab_widget.currentWidget()
        if not isinstance(current, ExerciseTab):
            return
            
        reply = QMessageBox.question(
            self,
            "אישור ניקוי",
            f"האם אתה בטוח שברצונך למחוק את כל הנתונים מהעמוד '{current.exercise_name}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                # שימוש בפורמט הקובץ החדש עם שם הפרופיל
                path = Path.cwd() / f"exercise_{current.profile_name}_{current.exercise_name}.json"
                if path.exists():
                    os.remove(path)
                # מחיקת קובץ ישן אם קיים
                old_path = Path.cwd() / f"exercise_state_{current.exercise_name}.json"
                if old_path.exists():
                    os.remove(old_path)
                
                # מחק את הטאב הנוכחי
                idx = self.tab_widget.currentIndex()
                self.tab_widget.removeTab(idx)
                current.deleteLater()

                # אם זה היה הטאב האחרון, הצג דיאלוג ליצירת תרגיל חדש
                if self.tab_widget.count() == 0:
                    title, ok = QInputDialog.getText(self, "תרגיל ראשון", "שם התרגיל:")
                    if ok and title.strip():
                        tab = ExerciseTab(title, self.current_profile_name)
                        self.tab_widget.addTab(tab, title)
                
                # עדכן את גיליון הסיכום
                self._update_summary_tab()

                self.statusBar().showMessage(f"נמחקו כל הנתונים מהעמוד '{current.exercise_name}'", 2000)
            except Exception as e:
                QMessageBox.warning(self, "שגיאה בניקוי", str(e))

    def _clear_all_tabs(self):
        reply = QMessageBox.question(
            self,
            "אישור ניקוי",
            "האם אתה בטוח שברצונך למחוק את כל הנתונים מכל העמודים?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                # מחק את כל הקבצים של הפרופיל הנוכחי
                profile_name = self.current_profile_name or "ברירת מחדל"
                for file in Path.cwd().glob(f"exercise_{profile_name}_*.json"):
                    try:
                        os.remove(file)
                    except Exception:
                        pass
                # מחיקת קבצים ישנים אם קיימים
                for file in Path.cwd().glob("exercise_state_*.json"):
                    try:
                        os.remove(file)
                    except Exception:
                        pass

                # סגור את כל הטאבים
                while self.tab_widget.count() > 0:
                    tab = self.tab_widget.widget(0)
                    self.tab_widget.removeTab(0)
                    if isinstance(tab, ExerciseTab):
                        tab.deleteLater()

                self.statusBar().showMessage("נמחקו כל הנתונים וכל העמודים", 2000)

                # הצג דיאלוג ליצירת תרגיל חדש
                title, ok = QInputDialog.getText(self, "תרגיל ראשון", "שם התרגיל:")
                if ok and title.strip():
                    tab = ExerciseTab(title)
                    self.tab_widget.addTab(tab, title)

            except Exception as e:
                QMessageBox.warning(self, "שגיאה בניקוי", str(e))

    def closeEvent(self, event):
        if self._closing:
            event.accept()
            return
            
        # בדיקה אם יש שינויים שלא נשמרו
        unsaved_tabs = []
        for i in range(self.tab_widget.count()):
            tab = self.tab_widget.widget(i)
            if isinstance(tab, ExerciseTab) and tab._has_unsaved_changes:
                unsaved_tabs.append(tab)
        
        if unsaved_tabs:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Icon.Question)
            msg.setWindowTitle("שינויים לא שמורים")
            if len(unsaved_tabs) == 1:
                msg.setText(f"יש שינויים שלא נשמרו בעמוד '{unsaved_tabs[0].exercise_name}'.\nהאם ברצונך לשמור לפני היציאה?")
            else:
                msg.setText("יש שינויים שלא נשמרו במספר עמודים.\nהאם ברצונך לשמור לפני היציאה?")
            msg.setStandardButtons(
                QMessageBox.StandardButton.Save | 
                QMessageBox.StandardButton.Discard | 
                QMessageBox.StandardButton.Cancel
            )
            msg.setDefaultButton(QMessageBox.StandardButton.Save)
            ret = msg.exec()
            
            if ret == QMessageBox.StandardButton.Save:
                # שמירת כל העמודים עם שינויים
                for tab in unsaved_tabs:
                    try:
                        tab.save_state()
                    except Exception:
                        pass
                self._closing = True
                event.accept()
            elif ret == QMessageBox.StandardButton.Discard:
                self._closing = True
                event.accept()
            else:  # Cancel
                event.ignore()
                return
        else:
            self._closing = True
            event.accept()


def apply_stylesheet(app: Any):
    # הגדרת סגנון כללי לאפליקציה
    app.setStyleSheet("""
        QMainWindow {
            background-color: #f0f0f0;
        }
        QTabWidget::pane {
            border: 1px solid #cccccc;
            background: white;
            border-radius: 5px;
        }
        QTabBar::tab {
            background: #e1e1e1;
            border: 1px solid #cccccc;
            padding: 8px 15px;
            margin-right: 2px;
            border-top-left-radius: 4px;
            border-top-right-radius: 4px;
            font-size: 11pt;
        }
        QTabBar::tab:selected {
            background: white;
            border-bottom-color: white;
        }
        QLineEdit {
            padding: 6px;
            border: 1px solid #cccccc;
            border-radius: 4px;
            background-color: white;
            font-size: 11pt;
        }
        QLineEdit:focus {
            border: 1px solid #2196F3;
        }
        QPushButton {
            padding: 6px 12px;
            background-color: #2196F3;
            color: white;
            border: none;
            border-radius: 4px;
            font-size: 11pt;
        }
        QPushButton:hover {
            background-color: #1976D2;
        }
        QPushButton:disabled {
            background-color: #BDBDBD;
        }
        QTableWidget {
            border: 1px solid #cccccc;
            border-radius: 4px;
            font-size: 11pt;
        }
        QTableWidget::item {
            padding: 4px;
            min-height: 24px;
        }
        QTableWidget {
            padding: 4px;
            min-height: 400px;
        }
        QTableWidget {
            gridline-color: #cccccc;
        }
        QTableWidget::item:selected {
            background-color: #E3F2FD;
            color: black;
        }
        QHeaderView::section {
            background-color: #f5f5f5;
            padding: 6px;
            border: 1px solid #cccccc;
            font-size: 11pt;
            font-weight: bold;
        }
        QLabel {
            font-size: 11pt;
        }
        QMenu {
            background-color: white;
            border: 1px solid #cccccc;
        }
        QMenu::item {
            padding: 6px 20px;
        }
        QMenu::item:selected {
            background-color: #E3F2FD;
        }
        QStatusBar {
            background-color: #f5f5f5;
            color: #333333;
            font-size: 10pt;
        }
        QToolBar {
            background-color: #f5f5f5;
            border-bottom: 1px solid #cccccc;
            spacing: 5px;
            padding: 5px;
        }
        QToolBar QToolButton {
            background-color: #2196F3;
            color: white;
            border-radius: 4px;
            padding: 5px 10px;
            font-size: 11pt;
        }
        QToolBar QToolButton:hover {
            background-color: #1976D2;
        }
        QMessageBox {
            font-size: 11pt;
        }
        QMessageBox QPushButton {
            min-width: 80px;
        }
    """)

if __name__ == "__main__":
    if not _HAS_QT:
        raise RuntimeError("PySide6 is required to run the GUI. Install requirements from requirements.txt")
    app = QApplication(sys.argv)
    apply_stylesheet(app)
    window = MainWindow()

    # חפש קבצי שמירה קיימים לפרופיל הנוכחי
    profile_name = window.current_profile_name or "ברירת מחדל"
    exercise_files = list(Path.cwd().glob(f"exercise_{profile_name}_*.json"))
    
    # אם אין קבצים לפרופיל הנוכחי, חפש קבצים ישנים (exercise_state_) ומיגרר אותם
    if not exercise_files:
        old_files = list(Path.cwd().glob("exercise_state_*.json"))
        if old_files and not window.current_profile_name:
            # מיגרציה של קבצים ישנים לפורמט החדש
            window.current_profile_name = "ברירת מחדל"
            profile_name = "ברירת מחדל"
            for old_file in old_files:
                old_name = old_file.stem.replace("exercise_state_", "")
                new_file = Path.cwd() / f"exercise_{profile_name}_{old_name}.json"
                try:
                    import shutil
                    shutil.copy2(old_file, new_file)
                except Exception:
                    pass
            exercise_files = list(Path.cwd().glob(f"exercise_{profile_name}_*.json"))
    
    if exercise_files:
        # אם יש קבצים קיימים, טען אותם
        for file in exercise_files:
            # חלץ את שם התרגיל מהקובץ
            exercise_name = file.stem.replace(f"exercise_{profile_name}_", "")
            tab = ExerciseTab(exercise_name, profile_name)
            window.tab_widget.addTab(tab, exercise_name)
    else:
        # אם אין קבצים קיימים, בקש שם תרגיל חדש
        title, ok = QInputDialog.getText(window, "תרגיל ראשון", "שם התרגיל:")
        if ok and title.strip():
            tab = ExerciseTab(title, profile_name)
            window.tab_widget.addTab(tab, title)

    window.show()
    app.exec()